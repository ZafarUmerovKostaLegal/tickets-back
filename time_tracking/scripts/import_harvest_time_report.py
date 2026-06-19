

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import json
import os
import re
import sys
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TT_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = TT_ROOT.parent
HARVEST_REPORT_BASENAME = "harvest_time_report_from2023-01-23to2026-06-02"
HARVEST_REPORT_BASENAME_FALLBACKS = ("harvest_time_report_from2023-01-23to2026-05-26",)
HARVEST_CSV_NAME = f"{HARVEST_REPORT_BASENAME}.csv"
HARVEST_XLSX_NAME = f"{HARVEST_REPORT_BASENAME}.xlsx"
DEFAULT_HARVEST_FILE = TT_ROOT / HARVEST_CSV_NAME
HARVEST_IMPORT_EMAIL_DOMAIN = "import.kostalegal.local"
HARVEST_IMPORT_AUTH_ID_FLOOR = 2_000_000_000
HARVEST_HOURS_QUANT = Decimal("0.01")
HARVEST_EXTRA_REQUIRED_TASKS = ("My mehnat registration",)

CHECKPOINT_VERSION = 1
PROJECT_KEY_SEP = "\x1f"


def _project_key(client_name: str, project_name: str) -> str:
    return f"{client_name}{PROJECT_KEY_SEP}{project_name}"


def _project_key_parts(key: str) -> tuple[str, str]:
    client, project = key.split(PROJECT_KEY_SEP, 1)
    return client, project


@dataclass(frozen=True)
class ProjectImportExpectation:
    total: Decimal
    billable: Decimal
    non_billable: Decimal
    row_count: int
    refs: frozenset[str]


def _expectation_from_project_rows(
    project_rows: list[HarvestRow],
    source_name: str,
) -> ProjectImportExpectation:
    total = Decimal("0")
    billable = Decimal("0")
    refs: set[str] = set()
    for row in project_rows:
        total += row.hours
        if row.is_billable:
            billable += row.hours
        refs.add(row.import_ref(source_name))
    total = _quantize_harvest_hours(total)
    billable = _quantize_harvest_hours(billable)
    return ProjectImportExpectation(
        total=total,
        billable=billable,
        non_billable=_quantize_harvest_hours(total - billable),
        row_count=len(project_rows),
        refs=frozenset(refs),
    )


def _build_csv_expectations_map(
    rows: list[HarvestRow],
    source_name: str,
    extra_project_pairs: list[tuple[str, str]] | None = None,
) -> dict[str, ProjectImportExpectation]:
    grouped: dict[str, list[HarvestRow]] = defaultdict(list)
    for row in rows:
        grouped[_project_key(row.client_name, row.project_name)].append(row)
    out = {
        key: _expectation_from_project_rows(proj_rows, source_name)
        for key, proj_rows in grouped.items()
    }
    empty = ProjectImportExpectation(
        total=Decimal("0"),
        billable=Decimal("0"),
        non_billable=Decimal("0"),
        row_count=0,
        refs=frozenset(),
    )
    for client_name, project_name in extra_project_pairs or []:
        out.setdefault(_project_key(client_name, project_name), empty)
    return out


def _project_db_matches_expectation(
    expected: ProjectImportExpectation,
    *,
    db_total: Decimal,
    db_billable: Decimal,
    db_non_billable: Decimal,
    db_row_count: int,
    db_refs: set[str],
) -> bool:
    return (
        _quantize_harvest_hours(db_total) == expected.total
        and _quantize_harvest_hours(db_billable) == expected.billable
        and _quantize_harvest_hours(db_non_billable) == expected.non_billable
        and db_row_count == expected.row_count
        and db_refs == set(expected.refs)
    )


async def _discover_complete_projects_in_db(
    session: Any,
    *,
    project_pairs: list[tuple[str, str]],
    csv_expectations: dict[str, ProjectImportExpectation],
    harvest_source_name: str,
) -> set[str]:
    """Проекты, у которых записи harvest-import:{file}:* в БД совпадают с CSV 1:1."""
    from sqlalchemy import case, func, select

    from infrastructure.models import (
        TimeEntryModel,
        TimeManagerClientModel,
        TimeManagerClientProjectModel,
    )

    harvest_prefix = f"harvest-import:{harvest_source_name}:"

    clients_by_norm: dict[str, tuple[str, str]] = {}
    client_id_to_name: dict[str, str] = {}
    client_result = await session.execute(select(TimeManagerClientModel))
    for client in client_result.scalars().all():
        clients_by_norm[_norm(client.name)] = (client.id, client.name)
        client_id_to_name[client.id] = client.name

    projects_by_client: dict[str, dict[str, str]] = defaultdict(dict)
    project_result = await session.execute(select(TimeManagerClientProjectModel))
    for project in project_result.scalars().all():
        projects_by_client[project.client_id][_norm(project.name)] = project.id

    pair_to_project_id: dict[str, str] = {}
    for client_name, project_name in project_pairs:
        client = clients_by_norm.get(_norm(client_name))
        if client is None:
            continue
        client_id, _ = client
        project_id = projects_by_client.get(client_id, {}).get(_norm(project_name))
        if project_id:
            pair_to_project_id[_project_key(client_name, project_name)] = project_id

    if not pair_to_project_id:
        return set()

    agg_result = await session.execute(
        select(
            TimeEntryModel.project_id,
            func.count(TimeEntryModel.id).label("rows"),
            func.coalesce(func.sum(TimeEntryModel.hours), 0).label("total"),
            func.coalesce(
                func.sum(
                    case((TimeEntryModel.is_billable.is_(True), TimeEntryModel.hours), else_=0)
                ),
                0,
            ).label("billable"),
            func.coalesce(
                func.sum(
                    case((TimeEntryModel.is_billable.is_(False), TimeEntryModel.hours), else_=0)
                ),
                0,
            ).label("non_billable"),
        )
        .where(
            TimeEntryModel.voided_at.is_(None),
            TimeEntryModel.external_reference_url.like(f"{harvest_prefix}%"),
            TimeEntryModel.project_id.in_(list(pair_to_project_id.values())),
        )
        .group_by(TimeEntryModel.project_id)
    )
    agg_by_project_id: dict[str, tuple[int, Decimal, Decimal, Decimal]] = {}
    for row in agg_result.all():
        agg_by_project_id[str(row.project_id)] = (
            int(row.rows or 0),
            Decimal(str(row.total)),
            Decimal(str(row.billable)),
            Decimal(str(row.non_billable)),
        )

    refs_by_project_id: dict[str, set[str]] = defaultdict(set)
    refs_result = await session.execute(
        select(TimeEntryModel.project_id, TimeEntryModel.external_reference_url).where(
            TimeEntryModel.voided_at.is_(None),
            TimeEntryModel.external_reference_url.like(f"{harvest_prefix}%"),
            TimeEntryModel.project_id.in_(list(pair_to_project_id.values())),
        )
    )
    for project_id, external_reference_url in refs_result.all():
        ref = (external_reference_url or "").strip()
        if ref:
            refs_by_project_id[str(project_id)].add(ref)

    complete: set[str] = set()
    for project_key, project_id in pair_to_project_id.items():
        expected = csv_expectations.get(project_key)
        if expected is None:
            continue
        agg = agg_by_project_id.get(project_id)
        if agg is None:
            continue
        row_count, db_total, db_billable, db_non_billable = agg
        if _project_db_matches_expectation(
            expected,
            db_total=db_total,
            db_billable=db_billable,
            db_non_billable=db_non_billable,
            db_row_count=row_count,
            db_refs=refs_by_project_id.get(project_id, set()),
        ):
            complete.add(project_key)
    return complete


def _default_checkpoint_path(source_file: Path) -> Path:
    return source_file.with_name(f"{source_file.name}.harvest-import.checkpoint.json")


def _file_fingerprint(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _load_checkpoint(
    checkpoint_path: Path,
    source_file: Path,
    *,
    reset: bool = False,
) -> dict[str, Any]:
    if reset and checkpoint_path.is_file():
        checkpoint_path.unlink()
    if not checkpoint_path.is_file():
        return {
            "version": CHECKPOINT_VERSION,
            "source_file": source_file.name,
            "source_fingerprint": _file_fingerprint(source_file),
            "completed_project_keys": [],
            "updated_at": None,
        }
    data = json.loads(checkpoint_path.read_text(encoding="utf-8"))
    if data.get("version") != CHECKPOINT_VERSION:
        raise SystemExit(
            f"Неподдерживаемая версия checkpoint {checkpoint_path} "
            f"(version={data.get('version')}). Удалите файл или используйте --reset-checkpoint."
        )
    if data.get("source_file") != source_file.name:
        raise SystemExit(
            f"Checkpoint {checkpoint_path} для другого файла "
            f"({data.get('source_file')!r}, ожидался {source_file.name!r}). "
            "Используйте --reset-checkpoint."
        )
    fp = _file_fingerprint(source_file)
    if data.get("source_fingerprint") and data.get("source_fingerprint") != fp:
        raise SystemExit(
            f"CSV изменился после последнего импорта (checkpoint {checkpoint_path}). "
            "Используйте --reset-checkpoint для начала заново."
        )
    data.setdefault("completed_project_keys", [])
    return data


def _save_checkpoint(checkpoint_path: Path, data: dict[str, Any]) -> None:
    data = dict(data)
    data["version"] = CHECKPOINT_VERSION
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    checkpoint_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _filter_project_pairs(
    pairs: list[tuple[str, str]],
    *,
    client_filter: str | None,
    project_filter: str | None,
    project_index: int | None,
) -> list[tuple[str, str]]:
    if project_index is not None:
        if project_index < 1 or project_index > len(pairs):
            raise ValueError(
                f"--project-index {project_index} вне диапазона 1..{len(pairs)}"
            )
        return [pairs[project_index - 1]]
    out = pairs
    if client_filter:
        cf = _norm(client_filter)
        out = [p for p in out if _norm(p[0]) == cf]
    if project_filter:
        pf = _norm(project_filter)
        out = [p for p in out if _norm(p[1]) == pf]
    return out


def _project_number(
    project_pair: tuple[str, str], all_project_pairs: list[tuple[str, str]]
) -> int:
    return all_project_pairs.index(project_pair) + 1


def _print_import_progress(
    checkpoint_path: Path,
    source_file: Path,
    all_project_pairs: list[tuple[str, str]],
    *,
    completed_keys: set[str] | None = None,
    source_label: str = "checkpoint",
) -> None:
    if completed_keys is None:
        data = _load_checkpoint(checkpoint_path, source_file, reset=False)
        completed = set(data.get("completed_project_keys") or [])
    else:
        completed = set(completed_keys)
    total = len(all_project_pairs)
    done = sum(1 for pair in all_project_pairs if _project_key(*pair) in completed)
    pending = total - done
    pct = (100.0 * done / total) if total else 100.0
    print(f"Файл: {source_file.name}")
    if completed_keys is None:
        print(f"Checkpoint: {checkpoint_path}")
    else:
        print(f"Источник прогресса: {source_label}")
    print(f"Проектов всего: {total}")
    print(f"Загружено: {done} ({pct:.1f}%)")
    print(f"Осталось: {pending}")
    if pending:
        for idx, pair in enumerate(all_project_pairs, 1):
            if _project_key(*pair) not in completed:
                c, p = pair
                print(f"\nСледующий проект: #{idx} — {c} / {p}")
                break
    if completed:
        print("\nПоследние 5 завершённых проектов:")
        for key in list(completed)[-5:]:
            c, p = _project_key_parts(key)
            print(f"  - {c} / {p}")


def _format_progress(done: int, total: int, client_name: str, project_name: str) -> str:
    pct = (100.0 * done / total) if total else 100.0
    return (
        f"Прогресс: {done}/{total} проектов ({pct:.1f}%) — "
        f"готово: {client_name} / {project_name}"
    )


def _glob_harvest_reports(directory: Path) -> list[Path]:
    """Любые harvest_time_report*.csv / *.xlsx в каталоге, новейшие первыми."""
    try:
        matches = [
            p
            for pattern in ("harvest_time_report*.csv", "harvest_time_report*.xlsx")
            for p in directory.glob(pattern)
            if p.is_file()
        ]
    except OSError:
        return []
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return matches


def _harvest_file_candidates(preferred: Path) -> list[Path]:
    seen: set[str] = set()
    out: list[Path] = []

    def add(p: Path) -> None:
        key = str(p)
        if key not in seen:
            seen.add(key)
            out.append(p)

    # Если --file не передан (preferred == зашитый дефолт), приоритет — самый свежий
    # harvest_time_report*.csv в папке сервиса (его «тянем» по умолчанию). Если --file
    # указан явно и существует, он будет первым.
    try:
        is_default = preferred.resolve() == DEFAULT_HARVEST_FILE.resolve()
    except OSError:
        is_default = False

    if not is_default:
        add(preferred)

    # Авто-поиск: любой harvest_time_report*.csv (напр. "harvest_time_report (1).csv")
    # в папке сервиса и типичных местах — берём самый свежий по дате изменения.
    search_dirs = [
        TT_ROOT,
        Path("/tmp"),
        REPO_ROOT / "timetrackinck",
        Path.cwd(),
        Path.cwd() / "timetrackinck",
    ]
    for directory in search_dirs:
        for match in _glob_harvest_reports(directory):
            add(match)

    if is_default:
        add(preferred)

    names = [HARVEST_CSV_NAME, HARVEST_XLSX_NAME]
    for base in HARVEST_REPORT_BASENAME_FALLBACKS:
        names.append(f"{base}.csv")
        names.append(f"{base}.xlsx")
    for name in names:
        add(TT_ROOT / name)
        add(Path("/tmp") / name)
        add(REPO_ROOT / "timetrackinck" / name)
        add(Path.cwd() / name)
        add(Path.cwd() / "timetrackinck" / name)
    return out


def _resolve_harvest_file(preferred: Path) -> Path | None:
    for candidate in _harvest_file_candidates(preferred):
        if candidate.is_file():
            resolved = candidate.resolve()
            if str(resolved) != str(preferred):
                print(f"Используется файл: {resolved}")
            return resolved
    return None


def _projects_catalog_candidates(preferred: Path, time_report: Path | None = None) -> list[Path]:
    seen: set[str] = set()
    out: list[Path] = []

    def add(p: Path) -> None:
        key = str(p)
        if key not in seen:
            seen.add(key)
            out.append(p)

    add(preferred)
    if time_report is not None:
        add(time_report.parent / preferred.name)
        if preferred.name != preferred.stem:
            add(time_report.parent / preferred.stem)
    for base in (Path("/app"), Path("/tmp"), TT_ROOT, Path.cwd(), REPO_ROOT / "timetrackinck"):
        add(base / preferred.name)
        if preferred.suffix:
            add(base / preferred.stem)
    return out


def _resolve_projects_catalog_file(
    preferred: Path,
    *,
    time_report: Path | None = None,
) -> Path | None:
    for candidate in _projects_catalog_candidates(preferred, time_report):
        if candidate.is_file():
            resolved = candidate.resolve()
            if str(resolved) != str(preferred.resolve() if preferred.exists() else preferred):
                print(f"Используется каталог проектов: {resolved}")
            return resolved
    return None


def _make_async_url(url: str) -> str:
    u = (url or "").strip()
    if u.startswith("postgresql://"):
        return u.replace("postgresql://", "postgresql+asyncpg://", 1)
    return u


def _resolve_database_url(cli_url: str | None) -> str:
    if cli_url and cli_url.strip():
        return cli_url.strip()
    for key in ("TIME_TRACKING_DATABASE_URL", "DATABASE_URL"):
        val = (os.environ.get(key) or "").strip()
        if val:
            print(f"Подключение: env {key}")
            return val
    raise SystemExit(
        "Задайте URL PostgreSQL time tracking:\n"
        "  export TIME_TRACKING_DATABASE_URL='postgresql://user:pass@host:5432/kosta_time_tracking'\n"
        "или: --database-url postgresql://..."
    )


def _configure_database_url(database_url: str) -> None:
    os.environ["DATABASE_URL"] = database_url
    os.environ["TIME_TRACKING_DATABASE_URL"] = database_url
    if TT_ROOT.as_posix() not in sys.path:
        sys.path.insert(0, TT_ROOT.as_posix())
    try:
        from infrastructure.config import get_settings

        get_settings.cache_clear()
    except Exception:
        pass


def _configure_stdio_utf8() -> None:
    """Avoid argparse/help crashes on Windows non-UTF consoles."""
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream is None:
            continue
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def _norm(s: str | None) -> str:
    return " ".join((s or "").strip().lower().split())


def _norm_project_code(code: str | None) -> str | None:
    text = (code or "").strip()
    return text.lower() if text else None


def _resolve_harvest_project_code(
    harvest_code: str | None,
    *,
    existing_code_owner_name: str | None,
    new_project_name: str,
) -> str | None:
    """Код проекта в TT уникален в рамках клиента; Harvest иногда дублирует код между проектами."""
    code = (harvest_code or "").strip() or None
    if not code:
        return None
    if existing_code_owner_name is None:
        return code
    if _norm(existing_code_owner_name) == _norm(new_project_name):
        return code
    return None


def _billable_default_for_harvest_task(task_key: str, project_rows: list[HarvestRow]) -> bool:
    """Как в Harvest: задача non-billable только если все часы по ней non-billable."""
    task_rows = [r for r in project_rows if _norm(r.task_name) == task_key]
    if not task_rows:
        return True
    bill_h = sum((r.hours for r in task_rows if r.is_billable), Decimal("0"))
    non_h = sum((r.hours for r in task_rows if not r.is_billable), Decimal("0"))
    if bill_h <= 0 and non_h > 0:
        return False
    return True


def _expected_hours_by_harvest_user(
    all_rows: list[HarvestRow],
) -> dict[str, tuple[str, Decimal, Decimal, Decimal, int]]:
    acc: dict[str, list] = {}
    for r in all_rows:
        key = r.harvest_user_key
        if key not in acc:
            acc[key] = [_harvest_display_name(r), Decimal("0"), Decimal("0"), Decimal("0"), 0]
        acc[key][1] += r.hours
        if r.is_billable:
            acc[key][2] += r.hours
        else:
            acc[key][3] += r.hours
        acc[key][4] += 1
    return {
        k: (str(v[0]), v[1], v[2], v[3], int(v[4]))
        for k, v in acc.items()
    }


def _expected_task_hours_map(project_rows: list[HarvestRow]) -> dict[str, tuple[Decimal, Decimal, Decimal]]:
    acc: dict[str, list[Decimal]] = defaultdict(
        lambda: [Decimal("0"), Decimal("0"), Decimal("0")]
    )
    for r in project_rows:
        key = _norm(r.task_name)
        acc[key][0] += r.hours
        if r.is_billable:
            acc[key][1] += r.hours
        else:
            acc[key][2] += r.hours
    for name in HARVEST_EXTRA_REQUIRED_TASKS:
        acc.setdefault(_norm(name), [Decimal("0"), Decimal("0"), Decimal("0")])
    return {k: (v[0], v[1], v[2]) for k, v in acc.items()}


def _parse_currency(raw: object) -> str:
    text = str(raw or "").strip().upper()
    for code in ("EUR", "USD", "UZS", "GBP", "RUB"):
        if code in text:
            return code
    return "EUR"


def _normalize_harvest_project_status(raw: object) -> str | None:
    text = str(raw or "").strip().lower()
    if not text:
        return None
    if any(token in text for token in ("archiv", "archive", "closed", "закрыт")):
        return "archived"
    if any(token in text for token in ("pause", "paused", "on hold", "на пауз")):
        return "paused"
    if any(token in text for token in ("budgeted", "budget", "бюджет")):
        return "budgeted"
    if any(token in text for token in ("active", "актив")):
        return "active"
    return None


def _harvest_project_is_archived(raw: object) -> bool | None:
    status = _normalize_harvest_project_status(raw)
    if status is None:
        return None
    return status == "archived"


def _status_from_catalog_row(row: dict[str, object]) -> tuple[str | None, bool | None]:
    status_candidates = (
        "Project Status",
        "Status",
        "Project State",
        "State",
        "Status Group",
        "Group",
        "Project Group",
    )
    for col in status_candidates:
        if col in row:
            raw = row.get(col)
            return _normalize_harvest_project_status(raw), _harvest_project_is_archived(raw)
    return None, None


def _parse_billable(raw: object) -> bool:
    s = str(raw or "").strip().lower()
    if s in ("no", "false", "0", "n", "non-billable", "non billable", "nonbillable"):
        return False
    if s in ("yes", "true", "1", "y", "billable"):
        return True
    return False


def _parse_work_date(raw: object) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _quantize_harvest_hours(h: Decimal) -> Decimal:
    return h.quantize(HARVEST_HOURS_QUANT, rounding=ROUND_HALF_UP)


def _clean_numeric_text(raw: object) -> str:
    # Harvest экспортирует числа с разделителем тысяч запятой и точкой-десятичным
    # (например "2,300,000.0" для UZS). Убираем запятые-разделители тысяч.
    return str(raw if raw is not None else "").strip().replace(",", "")


def _parse_hours(raw: object) -> Decimal | None:
    text = _clean_numeric_text(raw)
    if not text:
        return None
    try:
        h = Decimal(text)
    except Exception:
        return None
    if h <= 0:
        return None
    return _quantize_harvest_hours(h)


def _parse_money_rate(raw: object) -> Decimal | None:
    text = _clean_numeric_text(raw)
    if not text:
        return None
    try:
        val = Decimal(text)
    except Exception:
        return None
    return val.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def _harvest_user_rate_intervals(
    user_rows: list[HarvestRow],
    *,
    rate_attr: str = "billable_rate",
    billable_only: bool = True,
) -> list[tuple[Decimal, str, date, date]]:
    """Непересекающиеся интервалы ставок по датам работ из Harvest."""
    by_date: dict[date, tuple[Decimal, str]] = {}
    for r in user_rows:
        if billable_only and not r.is_billable:
            continue
        rate = getattr(r, rate_attr, None)
        if rate is None or rate <= 0:
            continue
        by_date[r.work_date] = (rate, r.currency)
    if not by_date:
        return []
    dates = sorted(by_date)
    intervals: list[tuple[Decimal, str, date, date]] = []
    start = dates[0]
    cur_amt, cur_cur = by_date[start]
    prev = start
    for work_date in dates[1:]:
        amt, cur = by_date[work_date]
        if amt == cur_amt and cur == cur_cur:
            prev = work_date
            continue
        intervals.append((cur_amt, cur_cur, start, prev))
        start = work_date
        prev = work_date
        cur_amt, cur_cur = amt, cur
    intervals.append((cur_amt, cur_cur, start, prev))
    return intervals


def _harvest_users_for_project(
    all_rows: list[HarvestRow],
    client_name: str,
    project_name: str,
) -> dict[str, HarvestRow]:
    ckey = _norm(client_name)
    pkey = _norm(project_name)
    out: dict[str, HarvestRow] = {}
    for r in all_rows:
        if _norm(r.client_name) == ckey and _norm(r.project_name) == pkey:
            out.setdefault(r.harvest_user_key, r)
    return out


def _user_needs_billable_rate_from_csv(user_rows: list[HarvestRow]) -> bool:
    return any(
        r.is_billable and r.billable_rate is not None and r.billable_rate > 0
        for r in user_rows
    )


def _user_has_billable_rate_for_harvest_rows(
    user_rows: list[HarvestRow],
    billable_rates: list[Any],
    *,
    project_currency: str,
) -> bool:
    """Ставка в валюте проекта покрывает billable-даты из CSV (не date.today())."""
    if not _user_needs_billable_rate_from_csv(user_rows):
        return True
    from application.hourly_rate_logic import filter_rates_by_currency, pick_rate_for_date

    scoped_rates = filter_rates_by_currency(billable_rates, project_currency)
    if not scoped_rates:
        return False
    for r in user_rows:
        if not r.is_billable or r.billable_rate is None or r.billable_rate <= 0:
            continue
        if pick_rate_for_date(scoped_rates, r.work_date) is None:
            return False
    return True


def _harvest_seconds_for_hours(hours: Decimal) -> int:
    return int((_quantize_harvest_hours(hours) * Decimal(3600)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _norm_notes_for_key(s: str | None) -> str:
    """Нормализация заметки для контентного ключа: схлопнуть пробелы, убрать края."""
    return " ".join((s or "").split())


def _entry_content_key(
    *,
    work_date: object,
    auth_user_id: int,
    project_id: str | None,
    task_name: str | None,
    hours: object,
    is_billable: bool,
    notes: str | None,
) -> str:
    """Стабильный контентный ключ записи времени.

    Ключ = дата + пользователь + проект + имя задачи + часы + billable + заметка.
    Не зависит от имени файла Harvest, номера строки и id задачи (id может
    поменяться при слиянии дублей задач, имя — нет). Используется для
    мультимножественного дифа: добавляем только те строки CSV, которых ещё нет
    в БД (с учётом легитимных повторов). Существующие записи не трогаем.
    """
    h = hours if isinstance(hours, Decimal) else Decimal(str(hours or "0"))
    h = _quantize_harvest_hours(h)
    wd = work_date.isoformat() if hasattr(work_date, "isoformat") else str(work_date)
    parts = "|".join(
        [
            wd,
            str(int(auth_user_id)),
            (project_id or "").strip(),
            _norm(task_name or ""),
            format(h, "f"),
            "1" if is_billable else "0",
            _norm_notes_for_key(notes),
        ]
    )
    return hashlib.sha1(parts.encode("utf-8")).hexdigest()


@dataclass(frozen=True)
class HarvestRow:
    source_row_number: int
    work_date: date
    client_name: str
    project_name: str
    project_code: str | None
    task_name: str
    notes: str | None
    hours: Decimal
    is_billable: bool
    first_name: str
    last_name: str
    employee_id: str | None
    billable_rate: Decimal | None
    cost_rate: Decimal | None
    currency: str
    external_reference_url: str | None

    @property
    def harvest_user_key(self) -> str:
        return _norm(f"{self.first_name} {self.last_name}")

    def import_ref(self, file_name: str) -> str:
        return f"harvest-import:{file_name}:{self.source_row_number}"


def _sorted_project_pairs(rows: list[HarvestRow]) -> list[tuple[str, str]]:
    return sorted(
        {(r.client_name, r.project_name) for r in rows},
        key=lambda p: (_norm(p[0]), _norm(p[1])),
    )


@dataclass(frozen=True)
class HarvestProjectCatalogEntry:
    client_name: str
    project_name: str
    project_code: str | None
    currency: str
    status: str | None = None
    is_archived: bool | None = None


def _load_projects_catalog(path: Path) -> list[HarvestProjectCatalogEntry]:
    """CSV/XLSX со списком проектов Harvest (в т.ч. без записей времени).

    Колонки: Client, Project, Project Code, Currency (как в time report).
    """
    suffix = path.suffix.lower()
    out: list[HarvestProjectCatalogEntry] = []
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                client_name = _csv_field(row, "Client").strip()
                project_name = _csv_field(row, "Project").strip()
                if not client_name or not project_name:
                    continue
                status, is_archived = _status_from_catalog_row(row)
                out.append(
                    HarvestProjectCatalogEntry(
                        client_name=client_name,
                        project_name=project_name,
                        project_code=_optional_text(_csv_field(row, "Project Code")),
                        currency=_parse_currency(_csv_field(row, "Currency")),
                        status=status,
                        is_archived=is_archived,
                    )
                )
        return out
    if suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["Harvest"] if "Harvest" in wb.sheetnames else wb[wb.sheetnames[0]]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            header_idx: dict[str, int] = {}
            if header_row:
                for idx, cell in enumerate(header_row):
                    key = _norm(str(cell or ""))
                    if key:
                        header_idx[key] = idx

            def _idx(*names: str, fallback: int) -> int:
                for name in names:
                    i = header_idx.get(_norm(name))
                    if i is not None:
                        return i
                return fallback

            idx_client = _idx("Client", fallback=1)
            idx_project = _idx("Project", fallback=2)
            idx_code = _idx("Project Code", fallback=3)
            idx_currency = _idx("Currency", fallback=19)
            idx_status = _idx(
                "Project Status",
                "Status",
                "Project State",
                "State",
                "Status Group",
                "Group",
                "Project Group",
                fallback=-1,
            )
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                client_name = str(row[idx_client] if len(row) > idx_client else "" or "").strip()
                project_name = str(row[idx_project] if len(row) > idx_project else "" or "").strip()
                if not client_name or not project_name:
                    continue
                status_raw = row[idx_status] if idx_status >= 0 and len(row) > idx_status else None
                out.append(
                    HarvestProjectCatalogEntry(
                        client_name=client_name,
                        project_name=project_name,
                        project_code=_optional_text(
                            str(row[idx_code] if len(row) > idx_code else "" or "")
                        ),
                        currency=_parse_currency(row[idx_currency] if len(row) > idx_currency else None),
                        status=_normalize_harvest_project_status(status_raw),
                        is_archived=_harvest_project_is_archived(status_raw),
                    )
                )
        finally:
            wb.close()
        return out
    raise SystemExit(
        f"Неподдерживаемый формат каталога проектов: {path.suffix} (нужен .csv или .xlsx)"
    )


def _merge_project_pairs(
    time_rows: list[HarvestRow],
    catalog: list[HarvestProjectCatalogEntry],
) -> list[tuple[str, str]]:
    pairs = {(r.client_name, r.project_name) for r in time_rows}
    for entry in catalog:
        pairs.add((entry.client_name, entry.project_name))
    return sorted(pairs, key=lambda p: (_norm(p[0]), _norm(p[1])))


def _build_harvest_meta_maps(
    time_rows: list[HarvestRow],
    catalog: list[HarvestProjectCatalogEntry],
) -> tuple[dict[str, str], dict[str, str], dict[str, str | None]]:
    """client_norm -> currency, project_key -> currency, project_key -> code."""
    project_currency: dict[str, str] = {}
    project_code: dict[str, str | None] = {}
    client_counts: dict[str, Counter] = defaultdict(Counter)

    for row in time_rows:
        key = _project_key(row.client_name, row.project_name)
        project_currency[key] = row.currency
        if row.project_code:
            project_code.setdefault(key, row.project_code)
        client_counts[_norm(row.client_name)][row.currency] += 1

    for entry in catalog:
        key = _project_key(entry.client_name, entry.project_name)
        project_currency.setdefault(key, entry.currency)
        if entry.project_code:
            project_code.setdefault(key, entry.project_code)
        client_counts[_norm(entry.client_name)][entry.currency] += 1

    client_currency = {
        client_key: counts.most_common(1)[0][0]
        for client_key, counts in client_counts.items()
    }
    return client_currency, project_currency, project_code


def _build_harvest_project_archived_map(
    catalog: list[HarvestProjectCatalogEntry],
) -> dict[str, bool]:
    out: dict[str, bool] = {}
    for entry in catalog:
        if entry.is_archived is None:
            continue
        out[_project_key(entry.client_name, entry.project_name)] = bool(entry.is_archived)
    return out


def _write_projects_catalog_from_time_rows(
    rows: list[HarvestRow],
    output_path: Path,
) -> int:
    """Генерирует каталог проектов из time report (только проекты с часами)."""
    _, project_currency_map, project_code_map = _build_harvest_meta_maps(rows, [])
    pairs = _sorted_project_pairs(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=("Client", "Project", "Project Code", "Currency", "Project Status"),
        )
        writer.writeheader()
        for client_name, project_name in pairs:
            key = _project_key(client_name, project_name)
            writer.writerow(
                {
                    "Client": client_name,
                    "Project": project_name,
                    "Project Code": project_code_map.get(key) or "",
                    "Currency": project_currency_map.get(key, "USD"),
                    "Project Status": "",
                }
            )
    return len(pairs)


def _csv_field(row: dict[str, str], name: str) -> str:
    if name in row:
        return str(row[name] or "").strip()
    target = name.strip().lower()
    for key, val in row.items():
        if key and key.strip().lower() == target:
            return str(val or "").strip()
    return ""


def _optional_text(raw: str) -> str | None:
    text = (raw or "").strip()
    return text or None


def _row_from_harvest_fields(
    *,
    source_row_number: int,
    work_date_raw: object,
    client_name: str,
    project_name: str,
    project_code_raw: object,
    task_name_raw: object,
    notes_raw: object,
    hours_raw: object,
    billable_raw: object,
    first_name: str,
    last_name: str,
    employee_id_raw: object,
    billable_rate_raw: object,
    cost_rate_raw: object,
    currency_raw: object,
    external_url_raw: object,
) -> HarvestRow | None:
    work_date = _parse_work_date(work_date_raw)
    client_name = (client_name or "").strip()
    project_name = (project_name or "").strip()
    if not work_date or not client_name or not project_name:
        return None
    hours = _parse_hours(hours_raw)
    if hours is None:
        return None
    first_name = (first_name or "").strip()
    last_name = (last_name or "").strip()
    if not first_name and not last_name:
        return None
    return HarvestRow(
        source_row_number=source_row_number,
        work_date=work_date,
        client_name=client_name,
        project_name=project_name,
        project_code=_optional_text(str(project_code_raw or "")),
        task_name=(str(task_name_raw or "Other research").strip() or "Other research"),
        notes=_optional_text(str(notes_raw or "")),
        hours=hours,
        is_billable=_parse_billable(billable_raw),
        first_name=first_name,
        last_name=last_name,
        employee_id=_optional_text(str(employee_id_raw or "")),
        billable_rate=_parse_money_rate(billable_rate_raw),
        cost_rate=_parse_money_rate(cost_rate_raw),
        currency=_parse_currency(currency_raw),
        external_reference_url=_optional_text(str(external_url_raw or "")),
    )


def _load_rows_from_csv(path: Path) -> list[HarvestRow]:
    out: list[HarvestRow] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for line_no, row in enumerate(reader, start=2):
            if not row or not any(str(v or "").strip() for v in row.values()):
                continue
            parsed = _row_from_harvest_fields(
                source_row_number=line_no,
                work_date_raw=_csv_field(row, "Date"),
                client_name=_csv_field(row, "Client"),
                project_name=_csv_field(row, "Project"),
                project_code_raw=_csv_field(row, "Project Code"),
                task_name_raw=_csv_field(row, "Task"),
                notes_raw=_csv_field(row, "Notes"),
                hours_raw=_csv_field(row, "Hours"),
                billable_raw=_csv_field(row, "Billable?"),
                first_name=_csv_field(row, "First Name"),
                last_name=_csv_field(row, "Last Name"),
                employee_id_raw=_csv_field(row, "Employee Id"),
                billable_rate_raw=_csv_field(row, "Billable Rate"),
                cost_rate_raw=_csv_field(row, "Cost Rate"),
                currency_raw=_csv_field(row, "Currency"),
                external_url_raw=_csv_field(row, "External Reference URL"),
            )
            if parsed is not None:
                out.append(parsed)
    return out


def _load_rows_from_xlsx(path: Path) -> list[HarvestRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Harvest"] if "Harvest" in wb.sheetnames else wb[wb.sheetnames[0]]
        out: list[HarvestRow] = []
        for line_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            parsed = _row_from_harvest_fields(
                source_row_number=line_no,
                work_date_raw=row[0],
                client_name=str(row[1] or ""),
                project_name=str(row[2] or ""),
                project_code_raw=row[3],
                task_name_raw=row[4],
                notes_raw=row[5],
                hours_raw=row[6],
                billable_raw=row[7],
                first_name=str(row[10] or ""),
                last_name=str(row[11] or ""),
                employee_id_raw=row[12] if len(row) > 12 else None,
                billable_rate_raw=row[15] if len(row) > 15 else None,
                cost_rate_raw=row[17] if len(row) > 17 else None,
                currency_raw=row[19] if len(row) > 19 else None,
                external_url_raw=row[20] if len(row) > 20 else None,
            )
            if parsed is not None:
                out.append(parsed)
        return out
    finally:
        wb.close()


def _load_rows(path: Path) -> list[HarvestRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_rows_from_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return _load_rows_from_xlsx(path)
    raise SystemExit(f"Неподдерживаемый формат файла: {path.suffix} (нужен .csv или .xlsx)")


def _build_name_index(
    *,
    display_name: str | None,
    email: str | None,
    auth_user_id: int,
    index: dict[str, int],
) -> None:
    if display_name:
        index[_norm(display_name)] = auth_user_id
    if email:
        local = email.split("@", 1)[0].replace(".", " ")
        index[_norm(local)] = auth_user_id


def _match_auth_user_id(row: HarvestRow, index: dict[str, int]) -> int | None:
    key = row.harvest_user_key
    if key in index:
        return index[key]
    rev = _norm(f"{row.last_name} {row.first_name}")
    if rev in index:
        return index[rev]
    return None


def _harvest_display_name(row: HarvestRow) -> str:
    return " ".join(p for p in (row.first_name.strip(), row.last_name.strip()) if p)


def _harvest_import_email(row: HarvestRow) -> str:
    slug = re.sub(r"[^a-z0-9]+", ".", row.harvest_user_key).strip(".") or "unknown"
    if row.employee_id:
        emp = re.sub(r"[^a-z0-9]+", "", str(row.employee_id).lower())
        if emp:
            slug = f"{slug}.{emp}"
    local = f"harvest.{slug}"
    max_local = 255 - len(HARVEST_IMPORT_EMAIL_DOMAIN) - 1
    if len(local) > max_local:
        local = local[:max_local]
    return f"{local}@{HARVEST_IMPORT_EMAIL_DOMAIN}"


async def _load_auth_users_for_import(auth_db_url: str) -> tuple[dict[str, int], dict[int, dict]]:
    from sqlalchemy import text
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from infrastructure.database import make_async_url

    engine = create_async_engine(make_async_url(auth_db_url), echo=False)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    index: dict[str, int] = {}
    by_id: dict[int, dict] = {}
    try:
        async with session_factory() as s:
            r = await s.execute(
                text(
                    """
                    SELECT id, email, display_name, picture, position, is_blocked, is_archived, time_tracking_role
                    FROM users
                    WHERE email IS NOT NULL AND trim(email) <> ''
                    ORDER BY id
                    """
                )
            )
            for row in r.mappings().all():
                auth_user_id = int(row["id"])
                email = str(row["email"]).strip()
                by_id[auth_user_id] = {
                    "auth_user_id": auth_user_id,
                    "email": email,
                    "display_name": row["display_name"],
                    "picture": row["picture"],
                    "position": row["position"],
                    "is_blocked": bool(row["is_blocked"]),
                    "is_archived": bool(row["is_archived"]),
                    "role": str(row["time_tracking_role"] or "").strip(),
                }
                _build_name_index(
                    display_name=str(row["display_name"]) if row["display_name"] is not None else None,
                    email=email,
                    auth_user_id=auth_user_id,
                    index=index,
                )
    finally:
        await engine.dispose()
    return index, by_id


async def _run(
    *,
    path: Path,
    execute: bool,
    database_url: str,
    auth_db_url: str = "",
    replace: bool = False,
    team_only: bool = False,
    batch_size: int = 0,
    checkpoint_file: Path | None = None,
    reset_checkpoint: bool = False,
    progress_only: bool = False,
    client_filter: str | None = None,
    project_filter: str | None = None,
    project_index: int | None = None,
    sync_from_db: bool = True,
    projects_file: Path | None = None,
) -> int:
    from sqlalchemy import and_, case, delete, func, select, update
    from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

    from application.client_expense_category_defaults import seed_default_expense_categories_for_client
    from application.hourly_rate_logic import filter_rates_by_currency, normalize_currency, pick_rate_for_date
    from application.project_billable_rate_sync import _delete_user_project_scoped_billable_rates
    from application.project_partner_requirement import (
        ensure_projects_have_partner_assignee,
        job_title_indicates_partner,
    )
    from infrastructure.repository_shared import _now_utc
    from infrastructure.models import (
        TimeEntryModel,
        TimeManagerClientModel,
        TimeManagerClientProjectModel,
        TimeManagerClientTaskModel,
        TimeTrackingUserModel,
    )
    from infrastructure.repository_access import UserProjectAccessRepository
    from infrastructure.repository_clients import ClientProjectRepository, ClientRepository, ClientTaskRepository
    from infrastructure.repository_entries import TimeEntryRepository
    from infrastructure.repository_rates import HourlyRateRepository, _rate_currency_key
    from infrastructure.repositories import TimeTrackingUserRepository

    rows = _load_rows(path)
    if not rows and not projects_file:
        print("Файл пуст или не содержит строк данных.")
        return 1

    catalog: list[HarvestProjectCatalogEntry] = []
    if projects_file is not None:
        catalog = _load_projects_catalog(projects_file)
        print(f"Каталог проектов: {projects_file} — {len(catalog)} проект(ов)")

    client_currency_map, project_currency_map, project_code_map = _build_harvest_meta_maps(
        rows, catalog
    )
    project_archived_map = _build_harvest_project_archived_map(catalog)
    all_project_pairs = _merge_project_pairs(rows, catalog)
    catalog_only_pairs = {
        (e.client_name, e.project_name)
        for e in catalog
        if not any(
            _norm(r.client_name) == _norm(e.client_name)
            and _norm(r.project_name) == _norm(e.project_name)
            for r in rows
        )
    }
    if catalog_only_pairs:
        print(
            f"Проектов только в каталоге (без часов в time report): {len(catalog_only_pairs)}"
        )

    print(f"Файл: {path}")
    print(f"Формат: {path.suffix.lower()} (1 строка файла = 1 запись времени)")
    harvest_source_name = path.name
    print(f"Клиентов: {len({p[0] for p in all_project_pairs})}")
    print(f"Проектов: {len(all_project_pairs)}")
    if rows:
        print(f"Строк в отчёте: {len(rows)}")
        print(f"Пользователей Harvest: {len({r.harvest_user_key for r in rows})}")
        expected_hours_total = sum((r.hours for r in rows), Decimal("0"))
        expected_billable_total = sum((r.hours for r in rows if r.is_billable), Decimal("0"))
        expected_non_billable_total = expected_hours_total - expected_billable_total
        print(
            f"Часы в файле Harvest: всего {expected_hours_total}, "
            f"billable {expected_billable_total}, non-billable {expected_non_billable_total}"
        )
    else:
        print("Строк времени в time report: 0 (только каталог проектов)")

    try:
        scope_project_pairs = _filter_project_pairs(
            all_project_pairs,
            client_filter=client_filter,
            project_filter=project_filter,
            project_index=project_index,
        )
    except ValueError as e:
        print(f"Ошибка фильтра проектов: {e}")
        return 1
    if client_filter or project_filter or project_index is not None:
        if len(scope_project_pairs) == 1:
            c, p = scope_project_pairs[0]
            n = _project_number((c, p), all_project_pairs)
            print(f"Фильтр: только проект #{n} — {c} / {p}")
        else:
            print(f"Фильтр: {len(scope_project_pairs)} проект(ов) из {len(all_project_pairs)}")

    checkpoint_path = checkpoint_file or _default_checkpoint_path(path)
    csv_expectations = _build_csv_expectations_map(
        rows, harvest_source_name, extra_project_pairs=catalog_only_pairs
    )

    checkpoint_data: dict[str, Any] | None = None
    completed_from_checkpoint: set[str] = set()
    if batch_size > 0:
        checkpoint_data = _load_checkpoint(checkpoint_path, path, reset=reset_checkpoint)
        completed_from_checkpoint = set(checkpoint_data.get("completed_project_keys") or [])

    engine = create_async_engine(_make_async_url(database_url), echo=False)
    session_factory = async_sessionmaker(
        engine,
        class_=AsyncSession,
        expire_on_commit=False,
        autocommit=False,
        autoflush=False,
    )

    completed_project_keys: set[str] = set(completed_from_checkpoint)
    if sync_from_db:
        if batch_size > 0:
            print("\nСверка CSV с БД (пропуск уже загруженных проектов)...")
        else:
            print("\nСверка CSV с БД (информативно; диф-режим обрабатывает все проекты)...")
        async with session_factory() as scan_session:
            completed_project_keys = await _discover_complete_projects_in_db(
                scan_session,
                project_pairs=all_project_pairs,
                csv_expectations=csv_expectations,
                harvest_source_name=harvest_source_name,
            )
        newly_found = completed_project_keys - completed_from_checkpoint
        print(
            f"В БД совпадают с CSV: {len(completed_project_keys)}/{len(all_project_pairs)} проектов"
        )
        if batch_size > 0 and newly_found:
            print(f"  (+{len(newly_found)} не были в checkpoint — checkpoint обновлён)")
        if batch_size > 0 or sync_from_db:
            if checkpoint_data is None:
                checkpoint_data = _load_checkpoint(checkpoint_path, path, reset=False)
            checkpoint_data["completed_project_keys"] = sorted(completed_project_keys)
            _save_checkpoint(checkpoint_path, checkpoint_data)
    elif batch_size > 0:
        completed_project_keys = completed_from_checkpoint

    if progress_only:
        if sync_from_db:
            _print_import_progress(
                checkpoint_path,
                path,
                all_project_pairs,
                completed_keys=completed_project_keys,
                source_label="БД + CSV",
            )
        else:
            _print_import_progress(checkpoint_path, path, all_project_pairs)
        await engine.dispose()
        return 0

    # В пакетном режиме (--batch-size) пропускаем уже завершённые проекты (резюм).
    # В обычном диф-режиме обрабатываем ВСЕ проекты: контентный диф сам пропустит
    # уже существующие строки и добавит недостающие. Это гарантирует, что ни одно
    # расхождение по часам не будет пропущено (агрегатная сверка discover могла
    # счесть проект «совпавшим», не проверяя отдельные строки).
    if batch_size > 0:
        pending_project_pairs = [
            pair
            for pair in scope_project_pairs
            if _project_key(*pair) not in completed_project_keys
        ]
    else:
        pending_project_pairs = list(scope_project_pairs)
    total_projects = len(all_project_pairs)
    done_count = sum(
        1 for pair in all_project_pairs if _project_key(*pair) in completed_project_keys
    )
    done_pct = (100.0 * done_count / total_projects) if total_projects else 100.0

    if batch_size > 0 or sync_from_db:
        print(
            f"\nК импорту: {len(pending_project_pairs)} проект(ов) "
            f"({sum(1 for r in rows if (r.client_name, r.project_name) in set(pending_project_pairs))} строк)"
        )
        print(
            f"Уже в БД: {done_count}/{total_projects} ({done_pct:.1f}%)"
        )
        if batch_size > 0:
            print(f"Пакетный режим: batch-size={batch_size}, checkpoint={checkpoint_path.name}")
        if execute and not pending_project_pairs:
            print("Все проекты из файла уже есть в БД. Для принудительного повтора: --replace")
            await engine.dispose()
            return 0
        if not execute:
            print(
                f"Dry-run: будет обработано {len(pending_project_pairs)} проектов "
                f"({sum(1 for r in rows if (r.client_name, r.project_name) in set(pending_project_pairs))} строк)."
            )

    import_batches: list[list[tuple[str, str]]] = []
    if batch_size > 0:
        import_batches = [
            pending_project_pairs[i : i + batch_size]
            for i in range(0, len(pending_project_pairs), batch_size)
        ]
    elif pending_project_pairs:
        import_batches = [pending_project_pairs]

    def build_user_index(users: list[TimeTrackingUserModel]) -> dict[str, int]:
        index: dict[str, int] = {}
        for u in users:
            _build_name_index(
                display_name=u.display_name,
                email=u.email,
                auth_user_id=int(u.auth_user_id),
                index=index,
            )
        return index

    match_auth_user_id = _match_auth_user_id

    async def find_tt_user_by_email(session: AsyncSession, email: str) -> TimeTrackingUserModel | None:
        r = await session.execute(
            select(TimeTrackingUserModel).where(TimeTrackingUserModel.email == email)
        )
        return r.scalars().one_or_none()

    async def next_harvest_auth_user_id(session: AsyncSession) -> int:
        from sqlalchemy import func

        r = await session.execute(select(func.max(TimeTrackingUserModel.auth_user_id)))
        mx = int(r.scalar_one_or_none() or 0)
        return max(HARVEST_IMPORT_AUTH_ID_FLOOR, mx + 1)

    def register_user_in_index(
        row: HarvestRow,
        *,
        auth_user_id: int,
        display_name: str,
        email: str,
        user_index: dict[str, int],
    ) -> None:
        user_index[row.harvest_user_key] = auth_user_id
        _build_name_index(
            display_name=display_name,
            email=email,
            auth_user_id=auth_user_id,
            index=user_index,
        )

    async def ensure_tt_user_for_harvest_import(session: AsyncSession, auth_user: dict) -> bool:
        """Создать или обновить TT-пользователя из auth; для Harvest всегда is_archived=True."""
        tur = TimeTrackingUserRepository(session)
        existing = await tur.get_by_auth_user_id(auth_user["auth_user_id"])
        pos = (str(auth_user.get("position") or "").strip()) or "Harvest import"
        await tur.upsert_user(
            auth_user_id=auth_user["auth_user_id"],
            email=auth_user["email"],
            display_name=auth_user.get("display_name"),
            picture=auth_user.get("picture"),
            role=str(auth_user.get("role") or ""),
            is_blocked=bool(auth_user.get("is_blocked", False)),
            is_archived=True,
            position=pos,
            update_position=existing is None,
        )
        await session.flush()
        return existing is None

    async def ensure_harvest_tt_user_archived(
        session: AsyncSession,
        auth_user_id: int,
        *,
        display_name: str | None,
        email: str | None,
    ) -> None:
        if not execute:
            return
        tur = TimeTrackingUserRepository(session)
        row = await tur.get_by_auth_user_id(auth_user_id)
        if row is None:
            return
        await tur.upsert_user(
            auth_user_id=auth_user_id,
            email=(email or row.email or "").strip(),
            display_name=display_name or row.display_name,
            picture=row.picture,
            role=row.role or "",
            is_blocked=bool(row.is_blocked),
            is_archived=True,
            position=row.position,
            update_position=False,
        )
        await session.flush()

    async def ensure_harvest_placeholder_user(
        session: AsyncSession,
        row: HarvestRow,
        *,
        user_index: dict[str, int],
        tt_by_auth: dict[int, TimeTrackingUserModel],
        harvest_placeholder_ids: dict[str, int],
        dry_run_placeholder_next: list[int],
    ) -> tuple[int, bool]:
        email = _harvest_import_email(row)
        display_name = _harvest_display_name(row)
        tur = TimeTrackingUserRepository(session)

        existing = await find_tt_user_by_email(session, email)
        if existing is not None:
            uid = int(existing.auth_user_id)
            if execute and not existing.is_archived:
                await tur.upsert_user(
                    auth_user_id=uid,
                    email=existing.email,
                    display_name=existing.display_name or display_name,
                    picture=existing.picture,
                    role=existing.role or "",
                    is_blocked=bool(existing.is_blocked),
                    is_archived=True,
                    position=existing.position,
                    update_position=False,
                )
                await session.flush()
                existing = await tur.get_by_auth_user_id(uid)
            register_user_in_index(
                row,
                auth_user_id=uid,
                display_name=display_name,
                email=email,
                user_index=user_index,
            )
            tt_by_auth[uid] = existing if existing is not None else await tur.get_by_auth_user_id(uid)
            harvest_placeholder_ids[row.harvest_user_key] = uid
            return uid, False

        cached = harvest_placeholder_ids.get(row.harvest_user_key)
        if cached is not None:
            return cached, False

        if not execute:
            dry_run_placeholder_next[0] -= 1
            uid = dry_run_placeholder_next[0]
            harvest_placeholder_ids[row.harvest_user_key] = uid
            register_user_in_index(
                row,
                auth_user_id=uid,
                display_name=display_name,
                email=email,
                user_index=user_index,
            )
            return uid, False

        uid = await next_harvest_auth_user_id(session)
        await tur.upsert_user(
            auth_user_id=uid,
            email=email,
            display_name=display_name,
            role="",
            is_blocked=False,
            is_archived=True,
            position="Harvest import",
            update_position=True,
        )
        await session.flush()
        refreshed = await tur.get_by_auth_user_id(uid)
        if refreshed is not None:
            tt_by_auth[uid] = refreshed
        harvest_placeholder_ids[row.harvest_user_key] = uid
        register_user_in_index(
            row,
            auth_user_id=uid,
            display_name=display_name,
            email=email,
            user_index=user_index,
        )
        return uid, True

    async def resolve_auth_user_id(
        session: AsyncSession,
        row: HarvestRow,
        user_index: dict[str, int],
        auth_index: dict[str, int],
        auth_by_id: dict[int, dict],
        tt_by_auth: dict[int, TimeTrackingUserModel],
        harvest_placeholder_ids: dict[str, int],
        dry_run_placeholder_next: list[int],
    ) -> tuple[int, bool, str]:
        """Всегда возвращает auth_user_id. Третье значение: tt|auth|harvest."""
        uid = match_auth_user_id(row, user_index)
        if uid is not None:
            return uid, False, "tt"

        if auth_index:
            uid = match_auth_user_id(row, auth_index)
            if uid is not None:
                auth_user = auth_by_id.get(uid)
                if auth_user is not None:
                    created = False
                    if execute:
                        created = await ensure_tt_user_for_harvest_import(session, auth_user)
                        register_user_in_index(
                            row,
                            auth_user_id=uid,
                            display_name=str(auth_user.get("display_name") or _harvest_display_name(row)),
                            email=auth_user["email"],
                            user_index=user_index,
                        )
                        refreshed = await TimeTrackingUserRepository(session).get_by_auth_user_id(uid)
                        if refreshed is not None:
                            tt_by_auth[uid] = refreshed
                    return uid, created, "auth"

        uid, created = await ensure_harvest_placeholder_user(
            session,
            row,
            user_index=user_index,
            tt_by_auth=tt_by_auth,
            harvest_placeholder_ids=harvest_placeholder_ids,
            dry_run_placeholder_next=dry_run_placeholder_next,
        )
        return uid, created, "harvest"

    async def find_first_partner_auth_user_id(session: AsyncSession) -> int | None:
        r = await session.execute(
            select(TimeTrackingUserModel.auth_user_id, TimeTrackingUserModel.position).order_by(
                TimeTrackingUserModel.is_archived.asc(),
                TimeTrackingUserModel.auth_user_id.asc(),
            )
        )
        for auth_user_id, position in r.all():
            if job_title_indicates_partner(position):
                return int(auth_user_id)
        return None

    async def finalize_imported_project(
        session: AsyncSession,
        *,
        client_id: str,
        project_id: str,
        granter: int | None,
        access_repo: UserProjectAccessRepository,
        project_repo: ClientProjectRepository,
        stats: Counter,
    ) -> None:
        row = await project_repo.get_by_id(client_id, project_id)
        if row is None:
            return
        try:
            await ensure_projects_have_partner_assignee(
                session,
                access_repo,
                {project_id},
                projects=project_repo,
            )
        except ValueError:
            partner_uid = await find_first_partner_auth_user_id(session)
            if partner_uid is None:
                print(
                    f"  ВНИМАНИЕ: проект «{row.name}» — нет партнёра в команде. "
                    "Добавьте партнёра вручную, иначе редактирование состава может не сохраниться."
                )
                return
            if granter is None:
                granter = partner_uid
            await access_repo.grant_access_if_absent(
                partner_uid,
                project_id,
                granted_by_auth_user_id=granter,
                projects=project_repo,
            )
            stats["project_partner_added"] += 1

    async def ensure_harvest_user_rates_from_csv(
        session: AsyncSession,
        *,
        harvest_user_key: str,
        sample_row: HarvestRow,
        auth_user_id: int,
        project_ids: list[str],
        stats: Counter,
        scoped_rows: list[HarvestRow] | None = None,
    ) -> None:
        user_rows = scoped_rows if scoped_rows is not None else [
            r for r in rows if r.harvest_user_key == harvest_user_key
        ]
        # Ставки берём по строкам ИМЕННО этого проекта (scoped_rows) и пишем их
        # с привязкой к проекту (applies_to_project_id=project_id). Так сумма
        # совпадает с Harvest для каждого проекта, даже если у сотрудника на
        # разных проектах разные ставки. Глобальные/ручные ставки НЕ трогаем.
        billable_intervals = _harvest_user_rate_intervals(user_rows)
        cost_intervals = _harvest_user_rate_intervals(
            user_rows, rate_attr="cost_rate", billable_only=False
        )
        if not billable_intervals and not cost_intervals:
            return

        hr = HourlyRateRepository(session)
        display = _harvest_display_name(sample_row)

        for project_id in project_ids:
            # Перезаписываем только harvest project-scoped ставки этого проекта
            # (идемпотентно для повторных прогонов); глобальные не удаляем.
            await _delete_user_project_scoped_billable_rates(session, auth_user_id, project_id)
            for row in await hr.list_by_user_and_kind(auth_user_id, "cost"):
                if getattr(row, "applies_to_project_id", None) == project_id:
                    await hr.delete(auth_user_id, row.id)

            for amt, cur, vf, vt in billable_intervals:
                try:
                    await hr.create(
                        auth_user_id=auth_user_id,
                        rate_kind="billable",
                        amount=amt,
                        currency=cur,
                        valid_from=vf,
                        valid_to=vt,
                        applies_to_project_id=project_id,
                    )
                    stats["hourly_rate_billable"] += 1
                except ValueError as e:
                    stats["hourly_rate_error"] += 1
                    print(f"  ОШИБКА billable-ставки {display}: {amt} {cur} ({vf}–{vt}): {e}")

            for amt, cur, vf, vt in cost_intervals:
                try:
                    await hr.create(
                        auth_user_id=auth_user_id,
                        rate_kind="cost",
                        amount=amt,
                        currency=cur,
                        valid_from=vf,
                        valid_to=vt,
                        applies_to_project_id=project_id,
                    )
                    stats["hourly_rate_cost"] += 1
                except ValueError as e:
                    stats["hourly_rate_error"] += 1
                    print(f"  ОШИБКА cost-ставки {display}: {amt} {cur} ({vf}–{vt}): {e}")

        if billable_intervals:
            parts = ", ".join(
                f"{amt} {normalize_currency(cur)} ({vf}–{vt})"
                for amt, cur, vf, vt in billable_intervals
            )
            print(f"  ставки {display}: {parts}")

    async def ensure_harvest_project_team_from_csv(
        session: AsyncSession,
        *,
        client_name: str,
        project_name: str,
        project_id: str,
        stats: Counter,
    ) -> None:
        """Все сотрудники из CSV: TT-пользователь, доступ к проекту, ставки."""
        nonlocal granter
        team = _harvest_users_for_project(rows, client_name, project_name)
        if not team:
            return
        if not execute:
            stats["project_team_members"] += len(team)
            return

        print(f"\n  Команда проекта «{project_name}» ({len(team)} чел. из CSV):")
        for hkey, sample in sorted(team.items(), key=lambda x: x[0]):
            uid, _created, user_source = await resolve_auth_user_id(
                session,
                sample,
                user_index,
                auth_index,
                auth_by_id,
                tt_by_auth,
                harvest_placeholder_ids,
                dry_run_placeholder_next,
            )
            if user_source == "harvest":
                harvest_only_names.add(hkey)

            email = (
                str(auth_by_id.get(uid, {}).get("email") or "").strip()
                if user_source == "auth"
                else _harvest_import_email(sample)
            )
            await ensure_harvest_tt_user_archived(
                session,
                uid,
                display_name=_harvest_display_name(sample),
                email=email or None,
            )
            refreshed = await TimeTrackingUserRepository(session).get_by_auth_user_id(uid)
            if refreshed is not None:
                tt_by_auth[uid] = refreshed
                if not refreshed.is_archived:
                    stats["entry_error"] += 1
                    print(
                        f"    ВНИМАНИЕ: {_harvest_display_name(sample)} "
                        f"(auth_user_id={uid}) не в архиве — доступ всё равно будет выдан"
                    )

            if granter is None:
                granter = uid

            access_key = (uid, project_id)
            user_proj_rows = [
                r for r in rows_for_project(client_name, project_name)
                if r.harvest_user_key == hkey
            ]
            await ensure_harvest_user_rates_from_csv(
                session,
                harvest_user_key=hkey,
                sample_row=sample,
                auth_user_id=uid,
                project_ids=[project_id],
                stats=stats,
                scoped_rows=user_proj_rows,
            )
            if access_key not in granted_access:
                await access_repo.grant_access_if_absent(
                    uid,
                    project_id,
                    granted_by_auth_user_id=granter,
                    projects=project_repo,
                )
                granted_access.add(access_key)
                stats["project_access_granted"] += 1

            stats["project_team_members"] += 1
            print(f"    + {_harvest_display_name(sample)}: список TT, доступ, ставки")
        await session.flush()

    async def ensure_all_csv_users_have_project_access(
        session: AsyncSession,
        stats: Counter,
    ) -> None:
        """Гарантировать доступ к проекту всем сотрудникам из CSV (идемпотентно)."""
        if not execute:
            return
        nonlocal granter
        for project_id, (_client_id, client_name, project_name) in project_meta.items():
            team = _harvest_users_for_project(rows, client_name, project_name)
            for hkey, sample in team.items():
                uid = harvest_user_auth_ids.get(hkey)
                if uid is None:
                    uid = _match_auth_user_id(sample, user_index)
                if uid is None:
                    uid = harvest_placeholder_ids.get(hkey)
                if uid is None:
                    stats["entry_error"] += 1
                    print(
                        f"  ОШИБКА доступа: {_harvest_display_name(sample)} — "
                        f"не найден auth_user_id"
                    )
                    continue
                if granter is None:
                    granter = uid
                access_key = (uid, project_id)
                if access_key in granted_access:
                    continue
                await access_repo.grant_access_if_absent(
                    uid,
                    project_id,
                    granted_by_auth_user_id=granter,
                    projects=project_repo,
                )
                granted_access.add(access_key)
                stats["project_access_granted"] += 1
        await session.flush()

    def expected_hours_for_project(client_name: str, project_name: str) -> Decimal:
        total = Decimal("0")
        ckey = _norm(client_name)
        pkey = _norm(project_name)
        for r in rows:
            if _norm(r.client_name) == ckey and _norm(r.project_name) == pkey:
                total += r.hours
        return total

    def expected_hours_breakdown_for_project(
        client_name: str, project_name: str
    ) -> tuple[Decimal, Decimal, Decimal]:
        total = Decimal("0")
        billable = Decimal("0")
        ckey = _norm(client_name)
        pkey = _norm(project_name)
        for r in rows:
            if _norm(r.client_name) == ckey and _norm(r.project_name) == pkey:
                total += r.hours
                if r.is_billable:
                    billable += r.hours
        return total, billable, total - billable

    async def find_client_by_name(session: AsyncSession, name: str) -> TimeManagerClientModel | None:
        target = _norm(name)
        r = await session.execute(select(TimeManagerClientModel))
        for c in r.scalars().all():
            if _norm(c.name) == target:
                return c
        return None

    async def find_project(
        session: AsyncSession,
        client_id: str,
        name: str,
    ) -> TimeManagerClientProjectModel | None:
        target = _norm(name)
        r = await session.execute(
            select(TimeManagerClientProjectModel).where(
                TimeManagerClientProjectModel.client_id == client_id
            )
        )
        for p in r.scalars().all():
            if _norm(p.name) == target:
                return p
        return None

    async def find_project_by_code(
        session: AsyncSession,
        client_id: str,
        code: str | None,
    ) -> TimeManagerClientProjectModel | None:
        target = _norm_project_code(code)
        if not target:
            return None
        r = await session.execute(
            select(TimeManagerClientProjectModel).where(
                TimeManagerClientProjectModel.client_id == client_id
            )
        )
        for project in r.scalars().all():
            if _norm_project_code(project.code) == target:
                return project
        return None

    async def harvest_project_code_for_create(
        session: AsyncSession,
        client_id: str,
        project_name: str,
        harvest_code: str | None,
    ) -> str | None:
        raw_code = (harvest_code or "").strip() or None
        if not raw_code:
            return None
        owner = await find_project_by_code(session, client_id, raw_code)
        resolved = _resolve_harvest_project_code(
            raw_code,
            existing_code_owner_name=owner.name if owner else None,
            new_project_name=project_name,
        )
        if resolved is None and owner is not None:
            print(
                f"  ВНИМАНИЕ: код «{raw_code}» уже у проекта «{owner.name}» — "
                f"для «{project_name}» код не задаём (Harvest дублирует Project Code)."
            )
        return resolved

    async def task_map_for_project(session: AsyncSession, project_id: str) -> dict[str, str]:
        r = await session.execute(
            select(TimeManagerClientTaskModel.name, TimeManagerClientTaskModel.id).where(
                TimeManagerClientTaskModel.project_id == project_id
            )
        )
        out: dict[str, str] = {}
        for name, tid in r.all():
            key = _norm(name)
            if key not in out:
                out[key] = str(tid)
        return out

    async def dedupe_project_tasks_by_name(
        session: AsyncSession,
        project_id: str,
        task_repo: ClientTaskRepository,
    ) -> int:
        r = await session.execute(
            select(TimeManagerClientTaskModel)
            .where(TimeManagerClientTaskModel.project_id == project_id)
            .order_by(TimeManagerClientTaskModel.created_at.asc())
        )
        groups: dict[str, list[TimeManagerClientTaskModel]] = defaultdict(list)
        for task in r.scalars().all():
            groups[_norm(task.name)].append(task)
        merged = 0
        for group in groups.values():
            if len(group) <= 1:
                continue
            keeper = group[0]
            for dup in group[1:]:
                await session.execute(
                    update(TimeEntryModel)
                    .where(TimeEntryModel.task_id == dup.id)
                    .values(task_id=keeper.id)
                )
                await task_repo.delete(project_id, dup.id)
                merged += 1
        return merged

    def rows_for_project(client_name: str, project_name: str) -> list[HarvestRow]:
        ckey = _norm(client_name)
        pkey = _norm(project_name)
        return [r for r in rows if _norm(r.client_name) == ckey and _norm(r.project_name) == pkey]

    async def ensure_harvest_project_tasks(
        session: AsyncSession,
        project_id: str,
        project_rows: list[HarvestRow],
        task_repo: ClientTaskRepository,
    ) -> dict[str, str]:
        if not execute:
            tmap: dict[str, str] = {}
            for r in project_rows:
                tmap[_norm(r.task_name)] = r.task_name
            for name in HARVEST_EXTRA_REQUIRED_TASKS:
                tmap.setdefault(_norm(name), name)
            return tmap

        merged = await dedupe_project_tasks_by_name(session, project_id, task_repo)
        if merged:
            print(f"  Объединены дубликаты задач проекта: {merged}")
        tmap = await task_map_for_project(session, project_id)

        display_names: dict[str, str] = {}
        for r in project_rows:
            display_names[_norm(r.task_name)] = r.task_name.strip()
        for name in HARVEST_EXTRA_REQUIRED_TASKS:
            display_names.setdefault(_norm(name), name)

        for key, display_name in sorted(display_names.items(), key=lambda x: x[1].lower()):
            billable_default = _billable_default_for_harvest_task(key, project_rows)
            task_id = tmap.get(key)
            if task_id:
                await task_repo.update(
                    project_id,
                    task_id,
                    {"billable_by_default": billable_default},
                )
            else:
                row_task = await task_repo.create(
                    project_id=project_id,
                    name=display_name,
                    default_billable_rate=Decimal("0"),
                    billable_by_default=billable_default,
                )
                task_id = row_task.id
                tmap[key] = task_id
                stats["task_created"] += 1
        await session.flush()
        return tmap

    async def db_task_breakdown_for_project(
        session: AsyncSession,
        project_id: str,
    ) -> dict[str, tuple[str, Decimal, Decimal, Decimal, bool]]:
        q = (
            select(
                TimeManagerClientTaskModel.name,
                func.coalesce(func.sum(TimeEntryModel.hours), 0).label("total"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(True), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("billable"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(False), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("non_billable"),
                TimeManagerClientTaskModel.billable_by_default,
            )
            .select_from(TimeManagerClientTaskModel)
            .outerjoin(
                TimeEntryModel,
                and_(
                    TimeEntryModel.task_id == TimeManagerClientTaskModel.id,
                    TimeEntryModel.voided_at.is_(None),
                ),
            )
            .where(TimeManagerClientTaskModel.project_id == project_id)
            .group_by(
                TimeManagerClientTaskModel.id,
                TimeManagerClientTaskModel.name,
                TimeManagerClientTaskModel.billable_by_default,
            )
        )
        r = await session.execute(q)
        out: dict[str, tuple[str, Decimal, Decimal, Decimal, bool]] = {}
        for row in r.all():
            key = _norm(row.name)
            if key in out:
                prev = out[key]
                out[key] = (
                    prev[0],
                    prev[1] + _quantize_harvest_hours(Decimal(str(row.total))),
                    prev[2] + _quantize_harvest_hours(Decimal(str(row.billable))),
                    prev[3] + _quantize_harvest_hours(Decimal(str(row.non_billable))),
                    bool(row.billable_by_default),
                )
            else:
                out[key] = (
                    str(row.name),
                    _quantize_harvest_hours(Decimal(str(row.total))),
                    _quantize_harvest_hours(Decimal(str(row.billable))),
                    _quantize_harvest_hours(Decimal(str(row.non_billable))),
                    bool(row.billable_by_default),
                )
        return out

    # Сверка учитывает ВСЕ harvest-записи проекта (и старые из прошлых файлов,
    # и новые контентные ref), т.к. диф добавляет недостающее к уже имеющемуся.
    harvest_file_prefix = "harvest-import:"

    async def aggregate_harvest_file_hours_for_project(
        session: AsyncSession,
        project_id: str,
    ) -> tuple[Decimal, Decimal, Decimal]:
        q = (
            select(
                func.coalesce(func.sum(TimeEntryModel.hours), 0).label("total"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(True), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("billable"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(False), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("non_billable"),
            )
            .where(
                TimeEntryModel.project_id == project_id,
                TimeEntryModel.voided_at.is_(None),
                TimeEntryModel.external_reference_url.like(f"{harvest_file_prefix}%"),
            )
        )
        row = (await session.execute(q)).one()
        return (
            _quantize_harvest_hours(Decimal(str(row.total))),
            _quantize_harvest_hours(Decimal(str(row.billable))),
            _quantize_harvest_hours(Decimal(str(row.non_billable))),
        )

    async def db_task_breakdown_harvest_file_for_project(
        session: AsyncSession,
        project_id: str,
    ) -> dict[str, tuple[str, Decimal, Decimal, Decimal, bool]]:
        q = (
            select(
                TimeManagerClientTaskModel.name,
                func.coalesce(func.sum(TimeEntryModel.hours), 0).label("total"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(True), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("billable"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(False), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("non_billable"),
                TimeManagerClientTaskModel.billable_by_default,
            )
            .select_from(TimeManagerClientTaskModel)
            .outerjoin(
                TimeEntryModel,
                and_(
                    TimeEntryModel.task_id == TimeManagerClientTaskModel.id,
                    TimeEntryModel.voided_at.is_(None),
                    TimeEntryModel.external_reference_url.like(f"{harvest_file_prefix}%"),
                ),
            )
            .where(TimeManagerClientTaskModel.project_id == project_id)
            .group_by(
                TimeManagerClientTaskModel.id,
                TimeManagerClientTaskModel.name,
                TimeManagerClientTaskModel.billable_by_default,
            )
        )
        return await _rows_to_task_breakdown(session, q)

    async def _rows_to_task_breakdown(session, q):
        r = await session.execute(q)
        out: dict[str, tuple[str, Decimal, Decimal, Decimal, bool]] = {}
        for row in r.all():
            key = _norm(str(row.name))
            if key in out:
                prev = out[key]
                out[key] = (
                    prev[0],
                    prev[1] + _quantize_harvest_hours(Decimal(str(row.total))),
                    prev[2] + _quantize_harvest_hours(Decimal(str(row.billable))),
                    prev[3] + _quantize_harvest_hours(Decimal(str(row.non_billable))),
                    bool(row.billable_by_default),
                )
            else:
                out[key] = (
                    str(row.name),
                    _quantize_harvest_hours(Decimal(str(row.total))),
                    _quantize_harvest_hours(Decimal(str(row.billable))),
                    _quantize_harvest_hours(Decimal(str(row.non_billable))),
                    bool(row.billable_by_default),
                )
        return out

    async def db_user_hours_harvest_file_for_projects(
        session: AsyncSession,
        project_ids: list[str],
    ) -> dict[int, tuple[Decimal, Decimal, Decimal, int]]:
        if not project_ids:
            return {}
        q = (
            select(
                TimeEntryModel.auth_user_id,
                func.coalesce(func.sum(TimeEntryModel.hours), 0).label("total"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(True), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("billable"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(False), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("non_billable"),
                func.count(TimeEntryModel.id).label("rows"),
            )
            .where(
                TimeEntryModel.project_id.in_(project_ids),
                TimeEntryModel.voided_at.is_(None),
                TimeEntryModel.external_reference_url.like(f"{harvest_file_prefix}%"),
            )
            .group_by(TimeEntryModel.auth_user_id)
        )
        r = await session.execute(q)
        out: dict[int, tuple[Decimal, Decimal, Decimal, int]] = {}
        for row in r.all():
            out[int(row.auth_user_id)] = (
                _quantize_harvest_hours(Decimal(str(row.total))),
                _quantize_harvest_hours(Decimal(str(row.billable))),
                _quantize_harvest_hours(Decimal(str(row.non_billable))),
                int(row.rows or 0),
            )
        return out

    async def db_user_hours_for_projects(
        session: AsyncSession,
        project_ids: list[str],
    ) -> dict[int, tuple[Decimal, Decimal, Decimal, int]]:
        if not project_ids:
            return {}
        q = (
            select(
                TimeEntryModel.auth_user_id,
                func.coalesce(func.sum(TimeEntryModel.hours), 0).label("total"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(True), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("billable"),
                func.coalesce(
                    func.sum(
                        case((TimeEntryModel.is_billable.is_(False), TimeEntryModel.hours), else_=0)
                    ),
                    0,
                ).label("non_billable"),
                func.count(TimeEntryModel.id).label("rows"),
            )
            .where(
                TimeEntryModel.project_id.in_(project_ids),
                TimeEntryModel.voided_at.is_(None),
            )
            .group_by(TimeEntryModel.auth_user_id)
        )
        r = await session.execute(q)
        out: dict[int, tuple[Decimal, Decimal, Decimal, int]] = {}
        for row in r.all():
            out[int(row.auth_user_id)] = (
                _quantize_harvest_hours(Decimal(str(row.total))),
                _quantize_harvest_hours(Decimal(str(row.billable))),
                _quantize_harvest_hours(Decimal(str(row.non_billable))),
                int(row.rows or 0),
            )
        return out

    async def db_user_hours_for_project(
        session: AsyncSession,
        project_id: str,
    ) -> dict[int, tuple[Decimal, Decimal, Decimal, int]]:
        return await db_user_hours_for_projects(session, [project_id])

    async def entry_exists_by_import_ref(session: AsyncSession, import_ref: str) -> bool:
        r = await session.execute(
            select(TimeEntryModel.id).where(
                TimeEntryModel.external_reference_url == import_ref,
                TimeEntryModel.voided_at.is_(None),
            ).limit(1)
        )
        return r.scalar_one_or_none() is not None

    auth_index: dict[str, int] = {}
    auth_by_id: dict[int, dict] = {}
    if auth_db_url.strip():
        print(f"Auth DB: сопоставление уволенных сотрудников по display_name")
        auth_index, auth_by_id = await _load_auth_users_for_import(auth_db_url.strip())

    try:
        async with session_factory() as session:
            tt_users = list((await session.execute(select(TimeTrackingUserModel))).scalars().all())
            user_index = build_user_index(tt_users)
            tt_by_auth = {int(u.auth_user_id): u for u in tt_users}
            harvest_placeholder_ids: dict[str, int] = {}
            dry_run_placeholder_next = [0]
            harvest_only_names: set[str] = set()
            harvest_user_sources: dict[str, str] = {}
            harvest_user_auth_ids: dict[str, int] = {}

            stats = Counter()
            client_repo = ClientRepository(session)
            project_repo = ClientProjectRepository(session)
            task_repo = ClientTaskRepository(session)
            entry_repo = TimeEntryRepository(session)
            access_repo = UserProjectAccessRepository(session)

            if execute and replace and batch_size <= 0:
                pairs = sorted({(r.client_name, r.project_name) for r in rows})
                for client_name, project_name in pairs:
                    existing_client = await find_client_by_name(session, client_name)
                    if existing_client is None:
                        continue
                    existing_project = await find_project(session, existing_client.id, project_name)
                    if existing_project is None:
                        continue
                    result = await session.execute(
                        delete(TimeEntryModel).where(TimeEntryModel.project_id == existing_project.id)
                    )
                    deleted = int(result.rowcount or 0)
                    if deleted:
                        print(
                            f"Удалены старые записи: {client_name} / {project_name} — {deleted} шт."
                        )
                await session.flush()

            # Режим дифа (по умолчанию, без --replace): существующие записи НЕ удаляем
            # и НЕ меняем. Дубли исключаются мультимножественным дифом по контентному
            # ключу (дата+пользователь+проект+задача+часы+billable+заметка), который не
            # зависит от имени файла отчёта и номера строки. Удаление доступно только
            # под явным --replace.
            if execute and not replace and batch_size <= 0:
                print(
                    "Режим дифа: существующие записи не трогаем; добавим только "
                    "отсутствующие в БД строки из отчёта Harvest."
                )

            client_cache: dict[str, str] = {}
            project_cache: dict[tuple[str, str], str] = {}
            project_meta: dict[str, tuple[str, str, str]] = {}
            task_cache: dict[str, dict[str, str]] = {}
            projects_tasks_initialized: set[str] = set()
            projects_team_initialized: set[str] = set()
            granted_access: set[tuple[int, str]] = set()
            granter: int | None = None

            async def _sync_harvest_currencies(
                client_id: str,
                client_name: str,
                project_id: str,
                project_name: str,
            ) -> None:
                ckey = _norm(client_name)
                pkey = _project_key(client_name, project_name)
                exp_client_cur = client_currency_map.get(ckey, "USD")
                exp_project_cur = project_currency_map.get(pkey, exp_client_cur)
                client_row = await client_repo.get_by_id(client_id)
                if client_row and (client_row.currency or "").strip().upper()[:10] != exp_client_cur:
                    await client_repo.update(client_id, {"currency": exp_client_cur})
                    stats["client_currency_synced"] += 1
                project_row = await project_repo.get_by_id(client_id, project_id)
                if project_row:
                    patch: dict[str, Any] = {}
                    if (project_row.currency or "").strip().upper()[:10] != exp_project_cur:
                        patch["currency"] = exp_project_cur
                    exp_archived = project_archived_map.get(pkey)
                    if exp_archived is not None and bool(project_row.is_archived) != exp_archived:
                        patch["is_archived"] = exp_archived
                    if patch:
                        await project_repo.update(client_id, project_id, patch)
                        if "currency" in patch:
                            stats["project_currency_synced"] += 1
                        if "is_archived" in patch:
                            stats["project_status_synced"] += 1

            async def _ensure_harvest_project_shell(
                client_name: str,
                project_name: str,
            ) -> str | None:
                """Создать клиента/проект без записей времени (из каталога Harvest)."""
                ckey = _norm(client_name)
                pkey_norm = _norm(project_name)
                pair_key = _project_key(client_name, project_name)
                exp_client_cur = client_currency_map.get(ckey, "USD")
                exp_project_cur = project_currency_map.get(pair_key, exp_client_cur)
                exp_code = project_code_map.get(pair_key)

                if ckey not in client_cache:
                    existing = await find_client_by_name(session, client_name)
                    if existing:
                        client_cache[ckey] = existing.id
                        stats["client_exists"] += 1
                    elif execute:
                        created = await client_repo.create(
                            name=client_name,
                            address=None,
                            currency=exp_client_cur,
                            invoice_due_mode="custom",
                            invoice_due_days_after_issue=30,
                            tax_percent=None,
                            tax2_percent=None,
                            discount_percent=None,
                        )
                        await seed_default_expense_categories_for_client(session, created.id)
                        client_cache[ckey] = created.id
                        stats["client_created"] += 1
                        await session.flush()
                    else:
                        return None
                client_id = client_cache[ckey]
                if client_id.startswith("<new:"):
                    return None

                cache_key = (client_id, pkey_norm)
                if cache_key not in project_cache:
                    existing_p = await find_project(session, client_id, project_name)
                    if existing_p:
                        project_cache[cache_key] = existing_p.id
                        project_meta[existing_p.id] = (client_id, client_name, project_name)
                        stats["project_exists"] += 1
                    elif execute:
                        project_code = await harvest_project_code_for_create(
                            session,
                            client_id,
                            project_name,
                            exp_code,
                        )
                        created_p = await project_repo.create(
                            client_id=client_id,
                            name=project_name,
                            code=project_code,
                            start_date=None,
                            end_date=None,
                            notes="Imported from Harvest (catalog, no time entries)",
                            report_visibility="managers_only",
                            project_type="time_and_materials",
                            currency=exp_project_cur,
                            billable_rate_type=None,
                            project_billable_rate_amount=None,
                            is_archived=project_archived_map.get(pair_key, False),
                        )
                        project_cache[cache_key] = created_p.id
                        project_meta[created_p.id] = (client_id, client_name, project_name)
                        stats["project_created"] += 1
                        await session.flush()
                    else:
                        return None
                project_id = project_cache[cache_key]
                await _sync_harvest_currencies(client_id, client_name, project_id, project_name)
                if project_id not in projects_tasks_initialized:
                    task_cache[project_id] = await ensure_harvest_project_tasks(
                        session,
                        project_id,
                        [],
                        task_repo,
                    )
                    projects_tasks_initialized.add(project_id)
                return project_id

            if execute and not team_only:
                cur_sync_clients_before = stats["client_currency_synced"]
                cur_sync_projects_before = stats["project_currency_synced"]
                for client_name, project_name in scope_project_pairs:
                    existing_client = await find_client_by_name(session, client_name)
                    if existing_client is None:
                        continue
                    existing_project = await find_project(
                        session, existing_client.id, project_name
                    )
                    if existing_project is None:
                        continue
                    await _sync_harvest_currencies(
                        existing_client.id,
                        client_name,
                        existing_project.id,
                        project_name,
                    )
                synced_clients = stats["client_currency_synced"] - cur_sync_clients_before
                synced_projects = stats["project_currency_synced"] - cur_sync_projects_before
                if synced_clients or synced_projects:
                    await session.flush()
                    print(
                        f"Синхронизация валют из Harvest: "
                        f"клиентов {synced_clients}, проектов {synced_projects}"
                    )

            # Сначала гарантируем TT-пользователя для каждого имени из Harvest (даже без регистрации в auth).
            pending_set = set(pending_project_pairs)
            user_setup_rows = [
                r
                for r in rows
                if (r.client_name, r.project_name) in pending_set
            ]
            unique_by_harvest_user: dict[str, HarvestRow] = {}
            for hr in user_setup_rows:
                unique_by_harvest_user.setdefault(hr.harvest_user_key, hr)
            print(f"\nПользователи Harvest (уникальных): {len(unique_by_harvest_user)}")
            for _key, sample in sorted(unique_by_harvest_user.items(), key=lambda x: x[0]):
                uid, tt_created, user_source = await resolve_auth_user_id(
                    session,
                    sample,
                    user_index,
                    auth_index,
                    auth_by_id,
                    tt_by_auth,
                    harvest_placeholder_ids,
                    dry_run_placeholder_next,
                )
                if execute:
                    email = (
                        str(auth_by_id.get(uid, {}).get("email") or "").strip()
                        if user_source == "auth"
                        else _harvest_import_email(sample)
                    )
                    await ensure_harvest_tt_user_archived(
                        session,
                        uid,
                        display_name=_harvest_display_name(sample),
                        email=email or None,
                    )
                    refreshed = await TimeTrackingUserRepository(session).get_by_auth_user_id(uid)
                    if refreshed is not None:
                        tt_by_auth[uid] = refreshed
                        if not refreshed.is_archived:
                            stats["entry_error"] += 1
                            print(
                                f"  ОШИБКА: пользователь {_harvest_display_name(sample)} "
                                f"(auth_user_id={uid}) не в архиве после импорта"
                            )
                if user_source == "harvest":
                    harvest_only_names.add(sample.harvest_user_key)
                    if tt_created and execute:
                        stats["tt_user_harvest_placeholder"] += 1
                        print(f"  + создан TT-пользователь Harvest: {_harvest_display_name(sample)}")
                elif user_source == "auth" and tt_created:
                    stats["tt_user_created"] += 1
                harvest_user_sources[_key] = user_source
                harvest_user_auth_ids[_key] = uid
                tt_row = tt_by_auth.get(uid)
                register_user_in_index(
                    sample,
                    auth_user_id=uid,
                    display_name=(
                        _harvest_display_name(sample)
                        if tt_row is None
                        else (tt_row.display_name or _harvest_display_name(sample))
                    ),
                    email=(
                        str(auth_by_id.get(uid, {}).get("email") or "").strip()
                        if user_source == "auth"
                        else (
                            (tt_row.email if tt_row else None)
                            or _harvest_import_email(sample)
                        )
                    ),
                    user_index=user_index,
                )

            total_projects = len(all_project_pairs)
            projects_done_checkpoint = len(completed_project_keys)

            async def _cleanup_projects_before_import(
                batch_pairs: list[tuple[str, str]],
            ) -> None:
                current_prefix = f"harvest-import:{harvest_source_name}:"
                for client_name, project_name in batch_pairs:
                    existing_client = await find_client_by_name(session, client_name)
                    if existing_client is None:
                        continue
                    existing_project = await find_project(
                        session, existing_client.id, project_name
                    )
                    if existing_project is None:
                        continue
                    if replace:
                        result = await session.execute(
                            delete(TimeEntryModel).where(
                                TimeEntryModel.project_id == existing_project.id
                            )
                        )
                        deleted = int(result.rowcount or 0)
                        if deleted:
                            print(
                                f"Удалены старые записи: {client_name} / {project_name} — {deleted} шт."
                            )
                    # Без --replace ничего не удаляем: диф по контентному ключу сам
                    # пропустит уже существующие записи и добавит только отсутствующие.
                await session.flush()

            # Контентный диф: подсчитываем уже существующие в БД harvest-записи
            # (по контентному ключу). Добавлять будем только недостающие строки CSV.
            db_content_counts: Counter = Counter()
            consumed_content_counts: Counter = Counter()
            if not team_only:
                task_name_by_id: dict[str, str] = {}
                for t in (
                    await session.execute(
                        select(TimeManagerClientTaskModel.id, TimeManagerClientTaskModel.name)
                    )
                ).all():
                    task_name_by_id[str(t.id)] = t.name or ""
                existing_rows = (
                    await session.execute(
                        select(
                            TimeEntryModel.work_date,
                            TimeEntryModel.auth_user_id,
                            TimeEntryModel.project_id,
                            TimeEntryModel.task_id,
                            TimeEntryModel.hours,
                            TimeEntryModel.is_billable,
                            TimeEntryModel.description,
                        ).where(
                            TimeEntryModel.voided_at.is_(None),
                            TimeEntryModel.external_reference_url.like("harvest-import:%"),
                        )
                    )
                ).all()
                for er in existing_rows:
                    db_content_counts[
                        _entry_content_key(
                            work_date=er.work_date,
                            auth_user_id=er.auth_user_id,
                            project_id=er.project_id,
                            task_name=task_name_by_id.get(str(er.task_id or ""), ""),
                            hours=er.hours,
                            is_billable=er.is_billable,
                            notes=er.description,
                        )
                    ] += 1
                print(
                    f"Контентный диф: в БД уже есть {sum(db_content_counts.values())} "
                    f"harvest-записей ({len(db_content_counts)} уникальных ключей)."
                )

            for batch_index, batch_pairs in enumerate(import_batches, 1):
                batch_set = set(batch_pairs)
                active_rows = [
                    r for r in rows if (r.client_name, r.project_name) in batch_set
                ]
                batch_project_meta: dict[str, tuple[str, str, str]] = {}

                if batch_size > 0 and execute:
                    first_client, first_project = batch_pairs[0]
                    global_no = _project_number(
                        (first_client, first_project), all_project_pairs
                    )
                    print(
                        f"\n=== Пакет {batch_index}/{len(import_batches)}: "
                        f"проект #{global_no}/{total_projects}: "
                        f"{first_client} / {first_project} "
                        f"({len(batch_pairs)} проект(ов), {len(active_rows)} строк) ==="
                    )
                    await _cleanup_projects_before_import(batch_pairs)

                for hr in active_rows:
                    ckey = _norm(hr.client_name)
                    if ckey not in client_cache:
                        existing = await find_client_by_name(session, hr.client_name)
                        if existing:
                            client_cache[ckey] = existing.id
                            stats["client_exists"] += 1
                            exp_client_cur = client_currency_map.get(ckey, hr.currency)
                            if execute and (existing.currency or "").strip().upper()[:10] != exp_client_cur:
                                await client_repo.update(existing.id, {"currency": exp_client_cur})
                                stats["client_currency_synced"] += 1
                        elif execute:
                            created = await client_repo.create(
                                name=hr.client_name,
                                address=None,
                                currency=client_currency_map.get(ckey, hr.currency),
                                invoice_due_mode="custom",
                                invoice_due_days_after_issue=30,
                                tax_percent=None,
                                tax2_percent=None,
                                discount_percent=None,
                            )
                            await seed_default_expense_categories_for_client(session, created.id)
                            client_cache[ckey] = created.id
                            stats["client_created"] += 1
                            await session.flush()
                        else:
                            client_cache[ckey] = f"<new:{hr.client_name}>"
                            stats["client_created"] += 1

                    client_id = client_cache[ckey]
                    if client_id.startswith("<new:"):
                        stats["entry_planned"] += 1
                        continue

                    pkey = (client_id, _norm(hr.project_name))
                    if pkey not in project_cache:
                        existing_p = await find_project(session, client_id, hr.project_name)
                        if existing_p:
                            project_cache[pkey] = existing_p.id
                            project_meta[existing_p.id] = (client_id, hr.client_name, hr.project_name)
                            stats["project_exists"] += 1
                            exp_project_cur = project_currency_map.get(
                                _project_key(hr.client_name, hr.project_name),
                                hr.currency,
                            )
                            if execute:
                                pair_key = _project_key(hr.client_name, hr.project_name)
                                patch: dict[str, Any] = {}
                                if (existing_p.currency or "").strip().upper()[:10] != exp_project_cur:
                                    patch["currency"] = exp_project_cur
                                exp_archived = project_archived_map.get(pair_key)
                                if exp_archived is not None and bool(existing_p.is_archived) != exp_archived:
                                    patch["is_archived"] = exp_archived
                                if patch:
                                    await project_repo.update(client_id, existing_p.id, patch)
                                    if "currency" in patch:
                                        stats["project_currency_synced"] += 1
                                    if "is_archived" in patch:
                                        stats["project_status_synced"] += 1
                        elif execute:
                            proj_dates = [
                                r.work_date
                                for r in rows
                                if _norm(r.client_name) == ckey and _norm(r.project_name) == _norm(hr.project_name)
                            ]
                            pair_key = _project_key(hr.client_name, hr.project_name)
                            project_code = await harvest_project_code_for_create(
                                session,
                                client_id,
                                hr.project_name,
                                project_code_map.get(pair_key, hr.project_code),
                            )
                            created_p = await project_repo.create(
                                client_id=client_id,
                                name=hr.project_name,
                                code=project_code,
                                start_date=min(proj_dates) if proj_dates else hr.work_date,
                                end_date=max(proj_dates) if proj_dates else None,
                                notes="Imported from Harvest",
                                report_visibility="managers_only",
                                project_type="time_and_materials",
                                currency=project_currency_map.get(pair_key, hr.currency),
                                billable_rate_type=None,
                                project_billable_rate_amount=None,
                                is_archived=project_archived_map.get(pair_key, False),
                            )
                            project_cache[pkey] = created_p.id
                            project_meta[created_p.id] = (client_id, hr.client_name, hr.project_name)
                            stats["project_created"] += 1
                            await session.flush()
                        else:
                            project_cache[pkey] = f"<new:{hr.project_name}>"
                            stats["project_created"] += 1

                    project_id = project_cache[pkey]
                    if project_id.startswith("<new:"):
                        stats["entry_planned"] += 1
                        continue

                    if project_id not in projects_tasks_initialized:
                        proj_rows = rows_for_project(hr.client_name, hr.project_name)
                        task_cache[project_id] = await ensure_harvest_project_tasks(
                            session,
                            project_id,
                            proj_rows,
                            task_repo,
                        )
                        projects_tasks_initialized.add(project_id)
                        if execute:
                            extra = ", ".join(HARVEST_EXTRA_REQUIRED_TASKS)
                            print(f"  Задачи проекта «{hr.project_name}»: из CSV + обязательные: {extra}")

                    if project_id not in projects_team_initialized:
                        await ensure_harvest_project_team_from_csv(
                            session,
                            client_name=hr.client_name,
                            project_name=hr.project_name,
                            project_id=project_id,
                            stats=stats,
                        )
                        projects_team_initialized.add(project_id)

                    tmap = task_cache[project_id]
                    tname = _norm(hr.task_name)
                    task_id = tmap.get(tname)
                    if not task_id:
                        stats["entry_error"] += 1
                        print(f"  ОШИБКА: нет задачи «{hr.task_name}» в проекте {hr.project_name}")
                        continue

                    auth_user_id, _tt_created, user_source = await resolve_auth_user_id(
                        session,
                        hr,
                        user_index,
                        auth_index,
                        auth_by_id,
                        tt_by_auth,
                        harvest_placeholder_ids,
                        dry_run_placeholder_next,
                    )
                    if user_source == "harvest":
                        harvest_only_names.add(hr.harvest_user_key)
                    if granter is None:
                        u = tt_by_auth.get(auth_user_id)
                        if u is None or not u.is_archived:
                            granter = auth_user_id
                    if auth_user_id not in tt_by_auth and execute:
                        refreshed = await TimeTrackingUserRepository(session).get_by_auth_user_id(auth_user_id)
                        if refreshed is not None:
                            tt_by_auth[auth_user_id] = refreshed

                    description = hr.notes
                    if not team_only:
                        # Мультимножественный диф: первые N одинаковых строк (N — сколько
                        # уже есть в БД) считаем существующими и пропускаем; легитимные
                        # повторы сверх этого числа добавляем. Существующее не трогаем.
                        content_key = _entry_content_key(
                            work_date=hr.work_date,
                            auth_user_id=auth_user_id,
                            project_id=project_id,
                            task_name=hr.task_name,
                            hours=hr.hours,
                            is_billable=hr.is_billable,
                            notes=description,
                        )
                        if (
                            db_content_counts.get(content_key, 0)
                            - consumed_content_counts[content_key]
                            > 0
                        ):
                            consumed_content_counts[content_key] += 1
                            stats["entry_duplicate"] += 1
                            continue
                        consumed_content_counts[content_key] += 1
                    if execute and not team_only:
                        if granter is None:
                            granter = auth_user_id
                        import_ref = f"harvest-import:content:{content_key}"
                        sec = _harvest_seconds_for_hours(hr.hours)
                        try:
                            entry_id = str(uuid.uuid4())
                            file_hours = hr.hours
                            session.add(
                                TimeEntryModel(
                                    id=entry_id,
                                    auth_user_id=auth_user_id,
                                    work_date=hr.work_date,
                                    duration_seconds=sec,
                                    hours=file_hours,
                                    rounded_hours=file_hours,
                                    is_billable=hr.is_billable,
                                    project_id=project_id,
                                    task_id=task_id,
                                    description=description,
                                    external_reference_url=import_ref,
                                    created_at=_now_utc(),
                                    updated_at=None,
                                )
                            )
                            stats["entry_created"] += 1
                        except ValueError as e:
                            stats["entry_error"] += 1
                            print(
                                f"  ОШИБКА записи {hr.work_date} {_harvest_display_name(hr)} "
                                f"({hr.hours} ч): {e}"
                            )
                    elif not execute:
                        stats["entry_planned"] += 1

                for shell_client, shell_project in batch_pairs:
                    if rows_for_project(shell_client, shell_project):
                        continue
                    shell_id = await _ensure_harvest_project_shell(shell_client, shell_project)
                    if shell_id and execute:
                        print(
                            f"  Каталог: проект без часов — {shell_client} / {shell_project}"
                        )

                if batch_size > 0 and execute:
                    scope_meta = {
                        pid: meta
                        for pid, meta in project_meta.items()
                        if (meta[1], meta[2]) in batch_set
                    }
                    if not scope_meta and active_rows:
                        print(
                            "ОШИБКА пакета: строки есть, но проект не создан в БД "
                            f"({batch_pairs[0][0]} / {batch_pairs[0][1]})."
                        )
                        await session.rollback()
                        return 1
                    if not scope_meta and not active_rows and batch_pairs:
                        print(
                            "ОШИБКА пакета: не удалось создать проект из каталога "
                            f"({batch_pairs[0][0]} / {batch_pairs[0][1]})."
                        )
                        await session.rollback()
                        return 1
                    for project_id, (client_id, _cn, _pn) in scope_meta.items():
                        await finalize_imported_project(
                            session,
                            client_id=client_id,
                            project_id=project_id,
                            granter=granter,
                            access_repo=access_repo,
                            project_repo=project_repo,
                            stats=stats,
                        )

                    await session.flush()

                    batch_hours_ok = True
                    batch_tasks_ok = True
                    batch_users_ok = True
                    batch_team_ok = True
                    batch_meta_ok = True
                    hr_repo = HourlyRateRepository(session)

                    for project_id, (_client_id, client_name, project_name) in scope_meta.items():
                        pkey = _project_key(client_name, project_name)
                        exp_client_cur = client_currency_map.get(_norm(client_name), "USD")
                        exp_project_cur = project_currency_map.get(pkey, exp_client_cur)
                        exp_archived = project_archived_map.get(pkey)
                        client_row = await client_repo.get_by_id(_client_id)
                        if client_row is None:
                            batch_meta_ok = False
                            print(
                                f"  [мета] {client_name} / {project_name}: "
                                "клиент не найден в БД"
                            )
                        elif (client_row.currency or "").strip().upper()[:10] != exp_client_cur:
                            batch_meta_ok = False
                            print(
                                f"  [мета] {client_name} / {project_name}: "
                                f"валюта клиента файл {exp_client_cur}, БД {client_row.currency}"
                            )
                        project_row = await project_repo.get_by_id(_client_id, project_id)
                        if project_row is None:
                            batch_meta_ok = False
                            print(
                                f"  [мета] {client_name} / {project_name}: "
                                "проект не найден в БД"
                            )
                        else:
                            if (project_row.currency or "").strip().upper()[:10] != exp_project_cur:
                                batch_meta_ok = False
                                print(
                                    f"  [мета] {client_name} / {project_name}: "
                                    f"валюта проекта файл {exp_project_cur}, БД {project_row.currency}"
                                )
                            if exp_archived is not None and bool(project_row.is_archived) != exp_archived:
                                batch_meta_ok = False
                                print(
                                    f"  [мета] {client_name} / {project_name}: "
                                    f"архивность файл {exp_archived}, БД {bool(project_row.is_archived)}"
                                )
                        db_total, db_billable, db_non_billable = await aggregate_harvest_file_hours_for_project(
                            session, project_id
                        )
                        exp_total, exp_billable, exp_non_billable = expected_hours_breakdown_for_project(
                            client_name, project_name
                        )
                        if not (
                            db_total == exp_total
                            and db_billable == exp_billable
                            and db_non_billable == exp_non_billable
                        ):
                            batch_hours_ok = False
                            print(
                                f"  [РАСХОЖДЕНИЕ] {client_name} / {project_name}: "
                                f"файл {exp_total}, БД {db_total}"
                            )

                        proj_rows = rows_for_project(client_name, project_name)
                        expected_tasks = _expected_task_hours_map(proj_rows)
                        db_tasks = await db_task_breakdown_harvest_file_for_project(session, project_id)
                        for key in expected_tasks:
                            exp_total_t, exp_bill_t, exp_non_t = expected_tasks[key]
                            exp_total_t = _quantize_harvest_hours(exp_total_t)
                            exp_bill_t = _quantize_harvest_hours(exp_bill_t)
                            exp_non_t = _quantize_harvest_hours(exp_non_t)
                            exp_billable_flag = _billable_default_for_harvest_task(key, proj_rows)
                            if key not in db_tasks:
                                if exp_total_t > 0:
                                    batch_tasks_ok = False
                                    print(
                                        f"  [задачи] {client_name} / {project_name}: "
                                        f"«{key}» нет в БД ({exp_total_t} ч в файле)"
                                    )
                                continue
                            _name, db_t, db_b, db_n, db_flag = db_tasks[key]
                            task_issues: list[str] = []
                            if db_t != exp_total_t:
                                task_issues.append(f"часы файл {exp_total_t}, БД {db_t}")
                            if db_b != exp_bill_t:
                                task_issues.append(f"billable файл {exp_bill_t}, БД {db_b}")
                            if db_n != exp_non_t:
                                task_issues.append(f"non-billable файл {exp_non_t}, БД {db_n}")
                            if exp_total_t > 0 and db_flag != exp_billable_flag:
                                task_issues.append(
                                    f"billable-флаг файл {exp_billable_flag}, БД {db_flag}"
                                )
                            if task_issues:
                                batch_tasks_ok = False
                                print(
                                    f"  [задачи] {client_name} / {project_name}: "
                                    f"«{_name}»: {'; '.join(task_issues)}"
                                )

                        team = _harvest_users_for_project(rows, client_name, project_name)
                        for hkey, sample in team.items():
                            uid = harvest_user_auth_ids.get(hkey) or _match_auth_user_id(sample, user_index)
                            if uid is None:
                                batch_team_ok = False
                                print(
                                    f"  [команда] {client_name} / {project_name}: "
                                    f"пользователь не найден — {_harvest_display_name(sample)}"
                                )
                                continue
                            user_proj_rows = [
                                r
                                for r in proj_rows
                                if r.harvest_user_key == hkey
                            ]
                            if not await access_repo.has_access(uid, project_id):
                                batch_team_ok = False
                                print(
                                    f"  [команда] {client_name} / {project_name}: "
                                    f"нет доступа — {_harvest_display_name(sample)}"
                                )
                            proj_cur = project_currency_map.get(
                                _project_key(client_name, project_name),
                                client_currency_map.get(_norm(client_name), "USD"),
                            )
                            if _user_needs_billable_rate_from_csv(user_proj_rows):
                                billable_rates = await hr_repo.list_by_user_and_kind(uid, "billable")
                                if not _user_has_billable_rate_for_harvest_rows(
                                    user_proj_rows,
                                    billable_rates,
                                    project_currency=proj_cur,
                                ):
                                    batch_team_ok = False
                                    print(
                                        f"  [команда] {client_name} / {project_name}: "
                                        f"нет billable-ставки в {proj_cur} — "
                                        f"{_harvest_display_name(sample)}"
                                    )

                    expected_users = _expected_hours_by_harvest_user(active_rows)
                    db_by_auth = await db_user_hours_harvest_file_for_projects(
                        session, list(scope_meta.keys())
                    )
                    for hkey, (name, exp_total, exp_bill, exp_non, exp_rows) in expected_users.items():
                        uid = harvest_user_auth_ids.get(hkey) or user_index.get(hkey)
                        if uid is None:
                            batch_users_ok = False
                            print(f"  [пользователи] не найден auth_user_id: {name}")
                            continue
                        db_total, db_bill, db_non, db_rows = db_by_auth.get(
                            uid, (Decimal("0"), Decimal("0"), Decimal("0"), 0)
                        )
                        user_issues: list[str] = []
                        if db_total != _quantize_harvest_hours(exp_total):
                            user_issues.append(
                                f"часы файл {_quantize_harvest_hours(exp_total)}, БД {db_total}"
                            )
                        if db_bill != _quantize_harvest_hours(exp_bill):
                            user_issues.append(
                                f"billable файл {_quantize_harvest_hours(exp_bill)}, БД {db_bill}"
                            )
                        if db_non != _quantize_harvest_hours(exp_non):
                            user_issues.append(
                                f"non-billable файл {_quantize_harvest_hours(exp_non)}, БД {db_non}"
                            )
                        if db_rows != exp_rows:
                            user_issues.append(f"строк файл {exp_rows}, БД {db_rows}")
                        if user_issues:
                            batch_users_ok = False
                            print(f"  [пользователи] {name}: {'; '.join(user_issues)}")

                    if not batch_hours_ok:
                        print("ОШИБКА пакета: часы по проектам не совпадают.")
                        print(
                            "Подсказка: проверьте строки «ОШИБКА записи» / «нет задачи» выше; "
                            "при повторном запуске уже загруженные проекты пропускаются (--progress)."
                        )
                        await session.rollback()
                        return 1
                    if not batch_tasks_ok:
                        print("ОШИБКА пакета: задачи не совпадают.")
                        await session.rollback()
                        return 1
                    if not batch_team_ok:
                        print("ОШИБКА пакета: команда/ставки не совпадают.")
                        await session.rollback()
                        return 1
                    if not batch_meta_ok:
                        print("ОШИБКА пакета: валюта/статус проекта не совпадают.")
                        await session.rollback()
                        return 1
                    if not batch_users_ok:
                        print("ОШИБКА пакета: пользователи/строки не совпадают.")
                        await session.rollback()
                        return 1

                    await session.commit()
                    assert checkpoint_data is not None
                    completed_list = list(checkpoint_data.get("completed_project_keys") or [])
                    completed_set = set(completed_list)
                    for client_name, project_name in batch_pairs:
                        completed_set.add(_project_key(client_name, project_name))
                    checkpoint_data["completed_project_keys"] = sorted(completed_set)
                    _save_checkpoint(checkpoint_path, checkpoint_data)
                    projects_done_checkpoint = len(completed_set)
                    last_client, last_project = batch_pairs[-1]
                    print(
                        _format_progress(
                            projects_done_checkpoint,
                            total_projects,
                            last_client,
                            last_project,
                        )
                    )

            if batch_size > 0:
                if execute:
                    print(
                        f"\nПакетный импорт завершён: {projects_done_checkpoint}/{total_projects} проектов "
                        f"({100.0 * projects_done_checkpoint / total_projects:.1f}%)."
                    )
                    print(
                        f"Записей времени всего: {stats['entry_created']} "
                        f"(дубликаты: {stats['entry_duplicate']}, ошибок: {stats['entry_error']})."
                    )
                else:
                    print("\nDry-run пакетного режима завершён. Для записи: --execute")
                if harvest_only_names:
                    print("\nСозданы/использованы TT-пользователи Harvest (нет в auth/TT):")
                    for name in sorted(harvest_only_names):
                        print(f"  - {name}")
                return 0

            if harvest_only_names:
                print("\nСозданы/использованы TT-пользователи Harvest (нет в auth/TT):")
                for name in sorted(harvest_only_names):
                    print(f"  - {name}")

            if execute:
                for project_id, (client_id, _client_name, _project_name) in project_meta.items():
                    await finalize_imported_project(
                        session,
                        client_id=client_id,
                        project_id=project_id,
                        granter=granter,
                        access_repo=access_repo,
                        project_repo=project_repo,
                        stats=stats,
                    )

                await ensure_all_csv_users_have_project_access(session, stats)

                imported_project_ids = list(project_meta.keys())
                print("\nСверка команды проекта (доступ + ставки из CSV):")
                team_ok = True
                hr_repo = HourlyRateRepository(session)
                for project_id, (_client_id, client_name, project_name) in project_meta.items():
                    team = _harvest_users_for_project(rows, client_name, project_name)
                    print(f"  Проект «{project_name}» ({len(team)} чел.):")
                    for hkey, sample in sorted(team.items(), key=lambda x: x[0]):
                        uid = harvest_user_auth_ids.get(hkey) or _match_auth_user_id(sample, user_index)
                        name = _harvest_display_name(sample)
                        user_proj_rows = [
                            r
                            for r in rows
                            if r.harvest_user_key == hkey
                            and _norm(r.client_name) == _norm(client_name)
                            and _norm(r.project_name) == _norm(project_name)
                        ]
                        if uid is None:
                            team_ok = False
                            print(f"    [ОТСУТСТВУЕТ] {name}: нет auth_user_id")
                            continue
                        has_access = await access_repo.has_access(uid, project_id)
                        needs_rate = _user_needs_billable_rate_from_csv(user_proj_rows)
                        has_rate = (
                            bool(await hr_repo.list_by_user_and_kind(uid, "billable"))
                            if needs_rate
                            else True
                        )
                        access_ok = has_access
                        rate_ok = has_rate
                        if not access_ok:
                            team_ok = False
                        if needs_rate and not rate_ok:
                            team_ok = False
                        status = "OK" if access_ok and rate_ok else "РАСХОЖДЕНИЕ"
                        rate_note = (
                            "ставка не требуется"
                            if not needs_rate
                            else ("ставка есть" if has_rate else "нет billable-ставки")
                        )
                        access_note = "доступ есть" if has_access else "нет доступа"
                        print(f"    [{status}] {name}: {access_note}, {rate_note}")

                await session.flush()

                if team_only:
                    if not team_ok:
                        print(
                            "\nВНИМАНИЕ: для части сотрудников остался [РАСХОЖДЕНИЕ] выше. "
                            "Остальные сохранены."
                        )
                    await session.commit()
                    print(
                        "\nГотово (режим --team-only): записи времени НЕ трогались.\n"
                        f"  TT-пользователей создано из auth: {stats['tt_user_created']}, "
                        f"архивных placeholder: {stats['tt_user_harvest_placeholder']};\n"
                        f"  доступ к проекту выдан: {stats['project_access_granted']}; "
                        f"партнёров добавлено: {stats['project_partner_added']};\n"
                        f"  ставок billable: {stats['hourly_rate_billable']}, "
                        f"cost: {stats['hourly_rate_cost']}."
                    )
                    return 0

                print(
                    f"\nЗаписей времени: добавлено {stats['entry_created']}, "
                    f"уже было в БД и пропущено {stats['entry_duplicate']}, "
                    f"ошибок {stats['entry_error']}"
                )

                print("\nСверка часов по проектам (Billable? из колонки 8 файла):")
                db_hours_total = Decimal("0")
                db_billable_total = Decimal("0")
                db_non_billable_total = Decimal("0")
                hours_ok = True
                for project_id, (client_id, client_name, project_name) in project_meta.items():
                    db_total, db_billable, db_non_billable = await entry_repo.aggregate_totals_for_project(
                        project_id
                    )
                    db_total = _quantize_harvest_hours(Decimal(str(db_total)))
                    db_billable = _quantize_harvest_hours(Decimal(str(db_billable)))
                    db_non_billable = _quantize_harvest_hours(Decimal(str(db_non_billable)))
                    exp_total, exp_billable, exp_non_billable = expected_hours_breakdown_for_project(
                        client_name, project_name
                    )
                    db_hours_total += db_total
                    db_billable_total += db_billable
                    db_non_billable_total += db_non_billable
                    project_ok = (
                        db_total == exp_total
                        and db_billable == exp_billable
                        and db_non_billable == exp_non_billable
                    )
                    if not project_ok:
                        hours_ok = False
                    status = "OK" if project_ok else "РАСХОЖДЕНИЕ"
                    print(
                        f"  [{status}] {project_name}: "
                        f"файл {exp_total} (billable {exp_billable}, non-billable {exp_non_billable}), "
                        f"БД {db_total} (billable {db_billable}, non-billable {db_non_billable})"
                    )

                print("\nСверка часов по задачам (как в Harvest):")
                tasks_ok = True
                for project_id, (_client_id, client_name, project_name) in project_meta.items():
                    proj_rows = rows_for_project(client_name, project_name)
                    expected_tasks = _expected_task_hours_map(proj_rows)
                    db_tasks = await db_task_breakdown_for_project(session, project_id)
                    print(f"  Проект «{project_name}»:")
                    for key in sorted(
                        expected_tasks,
                        key=lambda k: (db_tasks.get(k, (k,))[0] if k in db_tasks else k).lower(),
                    ):
                        exp_total, exp_bill, exp_non = expected_tasks[key]
                        exp_total = _quantize_harvest_hours(exp_total)
                        exp_bill = _quantize_harvest_hours(exp_bill)
                        exp_non = _quantize_harvest_hours(exp_non)
                        exp_billable_flag = _billable_default_for_harvest_task(key, proj_rows)
                        label = db_tasks[key][0] if key in db_tasks else key
                        if key not in db_tasks:
                            tasks_ok = False
                            print(f"    [ОТСУТСТВУЕТ] {label}: файл {exp_total} ч")
                            continue
                        name, db_total, db_bill, db_non, db_flag = db_tasks[key]
                        task_ok = (
                            db_total == exp_total
                            and db_bill == exp_bill
                            and db_non == exp_non
                            and db_flag == exp_billable_flag
                        )
                        if not task_ok:
                            tasks_ok = False
                        status = "OK" if task_ok else "РАСХОЖДЕНИЕ"
                        billable_label = "billable" if exp_billable_flag else "non-billable"
                        print(
                            f"    [{status}] {name} ({billable_label}): "
                            f"файл {exp_total} (b {exp_bill}, nb {exp_non}), "
                            f"БД {db_total} (b {db_bill}, nb {db_non}), "
                            f"флаг задачи={'billable' if db_flag else 'non-billable'}"
                        )

                print("\nСверка по пользователям (строки CSV = записи TT, часы 1:1):")
                users_ok = True
                expected_users = _expected_hours_by_harvest_user(rows)
                imported_project_ids = list(project_meta.keys())
                db_by_auth = await db_user_hours_for_projects(session, imported_project_ids)
                for hkey in sorted(expected_users, key=lambda k: expected_users[k][0].lower()):
                    name, exp_total, exp_bill, exp_non, exp_rows = expected_users[hkey]
                    uid = harvest_user_auth_ids.get(hkey) or user_index.get(hkey) or user_index.get(_norm(hkey))
                    if uid is None:
                        users_ok = False
                        print(f"  [ОТСУТСТВУЕТ] {name}: нет TT-пользователя")
                        continue
                    tt_row = tt_by_auth.get(uid)
                    if tt_row is None:
                        tt_row = await TimeTrackingUserRepository(session).get_by_auth_user_id(uid)
                    if tt_row is None or not tt_row.is_archived:
                        users_ok = False
                        print(f"  [НЕ В АРХИВЕ] {name} (auth_user_id={uid})")
                    db_total, db_bill, db_non, db_rows = db_by_auth.get(
                        uid, (Decimal("0"), Decimal("0"), Decimal("0"), 0)
                    )
                    user_ok = (
                        db_total == _quantize_harvest_hours(exp_total)
                        and db_bill == _quantize_harvest_hours(exp_bill)
                        and db_non == _quantize_harvest_hours(exp_non)
                        and db_rows == exp_rows
                    )
                    if not user_ok:
                        users_ok = False
                    status = "OK" if user_ok else "РАСХОЖДЕНИЕ"
                    print(
                        f"  [{status}] {name}: файл {exp_rows} строк / {exp_total} ч "
                        f"(b {exp_bill}, nb {exp_non}), "
                        f"БД {db_rows} строк / {db_total} ч (b {db_bill}, nb {db_non})"
                    )

                print(
                    f"\nИтого: файл {expected_hours_total} "
                    f"(billable {expected_billable_total}, non-billable {expected_non_billable_total}), "
                    f"БД {db_hours_total} (billable {db_billable_total}, non-billable {db_non_billable_total})"
                )
                # Жёсткая сверка «БД == файл 1:1» имеет смысл только в режиме --replace
                # (полная перезаливка проекта). В диф-режиме в БД легитимно может быть
                # больше: ручные записи, записи из прежних импортов, проекты, чья полная
                # история шире присланного отчёта. Поэтому здесь — предупреждения, без отката.
                strict = replace
                if not hours_ok or db_hours_total != expected_hours_total:
                    if strict:
                        print("ОШИБКА: сумма часов в БД не совпадает с файлом Harvest.")
                        await session.rollback()
                        return 1
                    print(
                        "ВНИМАНИЕ (диф): часы в БД ≠ суммы из файла — допустимо, "
                        "если в БД есть ручные/прежние записи. Откат не делаю."
                    )

                if not tasks_ok:
                    if strict:
                        print("ОШИБКА: часы или billable-флаг задач не совпадают с Harvest.")
                        await session.rollback()
                        return 1
                    print("ВНИМАНИЕ (диф): часы/флаги задач отличаются от файла — допустимо.")

                if not team_ok:
                    if strict:
                        print("ОШИБКА: не у всех сотрудников из CSV есть доступ к проекту и/или ставки.")
                        await session.rollback()
                        return 1
                    print(
                        "ВНИМАНИЕ (диф): не у всех сотрудников есть доступ/ставка — проверьте список выше."
                    )

                if not users_ok:
                    if strict:
                        print("ОШИБКА: пользователи или их часы/строки не совпадают с Harvest.")
                        await session.rollback()
                        return 1
                    print("ВНИМАНИЕ (диф): часы/строки пользователей отличаются от файла — допустимо.")

                if stats["hourly_rate_error"] > 0:
                    print(f"\nОШИБКА: не удалось выставить ставок: {stats['hourly_rate_error']}")
                    await session.rollback()
                    return 1

                if stats["entry_error"] > 0:
                    print(f"\nОШИБКА: не импортировано записей: {stats['entry_error']}")
                    await session.rollback()
                    return 1

                processed = stats["entry_created"] + stats["entry_duplicate"]
                if processed != len(rows):
                    print(
                        f"\nОШИБКА: в файле {len(rows)} строк, в БД создано/найдено {processed} "
                        f"(создано: {stats['entry_created']}, дубликаты: {stats['entry_duplicate']})."
                    )
                    await session.rollback()
                    return 1

                await session.commit()
                print("\nИмпорт выполнен.")
            else:
                print("\nDry-run (без записи в БД). Добавьте --execute для импорта.")

            print(
                f"Клиентов создано: {stats['client_created']}, уже было: {stats['client_exists']}; "
                f"проектов создано: {stats['project_created']}, уже было: {stats['project_exists']}; "
                f"статусов проектов синхронизировано: {stats['project_status_synced']}; "
                f"задач создано: {stats['task_created']}; "
                f"сотрудников в команде проекта: {stats['project_team_members']}; "
                f"доступ к проекту выдан: {stats['project_access_granted']}; "
                f"партнёров добавлено: {stats['project_partner_added']}; "
                f"TT из auth: {stats['tt_user_created']}, placeholder Harvest: {stats['tt_user_harvest_placeholder']}; "
                f"billable-ставок: {stats['hourly_rate_billable']}, cost-ставок: {stats['hourly_rate_cost']}; "
                f"записей времени: {stats['entry_created'] or stats['entry_planned']} "
                f"(уже в БД и пропущено: {stats['entry_duplicate']}, ошибок: {stats['entry_error']})."
            )
            expected = len(rows)
            if execute:
                imported = stats["entry_created"] + stats["entry_duplicate"]
            else:
                imported = stats["entry_planned"] + stats["entry_duplicate"]
            if imported + stats["entry_error"] < expected:
                print(
                    f"ВНИМАНИЕ: в отчёте {expected} строк, обработано {imported} "
                    f"(+ ошибок: {stats['entry_error']}). Часть строк не разобрана — проверьте лог."
                )
            elif execute:
                print(
                    f"Все {expected} строк отчёта учтены в БД "
                    f"(добавлено {stats['entry_created']}, уже было {stats['entry_duplicate']})."
                )
            else:
                print(
                    f"Dry-run диф: добавит {stats['entry_planned']} новых строк, "
                    f"пропустит {stats['entry_duplicate']} (уже в БД) из {expected}."
                )
    finally:
        await engine.dispose()
    return 0


def main() -> int:
    _configure_stdio_utf8()
    p = argparse.ArgumentParser(description="Импорт Harvest time report (.csv / .xlsx) в time tracking.")
    p.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_HARVEST_FILE,
        help=f"Путь к .csv или .xlsx (по умолчанию: {HARVEST_CSV_NAME}).",
    )
    p.add_argument(
        "--database-url",
        type=str,
        default="",
        help="PostgreSQL URL (иначе TIME_TRACKING_DATABASE_URL или DATABASE_URL).",
    )
    p.add_argument(
        "--auth-db-url",
        type=str,
        default="",
        help="Auth PostgreSQL URL для уволенных (env AUTH_DATABASE_URL).",
    )
    p.add_argument(
        "--replace",
        action="store_true",
        help="Удалить старые записи времени по проектам из файла перед импортом (чистый 1:1).",
    )
    p.add_argument(
        "--team-only",
        action="store_true",
        help=(
            "Только команда: завести пользователей (отсутствующих — архивными), "
            "выдать доступ к проекту и ставки. Записи времени НЕ трогаются, откатов нет. "
            "Запускать вместе с --execute."
        ),
    )
    p.add_argument(
        "--batch-size",
        type=int,
        default=0,
        metavar="N",
        help=(
            "Импортировать по N проектов за раз с commit и checkpoint после каждого пакета. "
            "N=1 — по одному проекту (можно прервать и продолжить). 0 — всё за один проход."
        ),
    )
    p.add_argument(
        "--checkpoint",
        type=Path,
        default=None,
        help="Путь к JSON checkpoint (по умолчанию: {csv}.harvest-import.checkpoint.json).",
    )
    p.add_argument(
        "--reset-checkpoint",
        action="store_true",
        help="Удалить checkpoint и начать импорт проектов заново.",
    )
    p.add_argument(
        "--progress",
        action="store_true",
        help="Показать прогресс по checkpoint и выйти (без импорта).",
    )
    p.add_argument("--client", type=str, default="", help="Импорт только указанного клиента.")
    p.add_argument("--project", type=str, default="", help="Импорт только указанного проекта.")
    p.add_argument(
        "--project-index",
        type=int,
        default=0,
        metavar="N",
        help="Импорт только проекта № N (1..749, как в sorted-списке).",
    )
    p.add_argument(
        "--no-sync-from-db",
        action="store_true",
        help="Не сверять CSV с БД при старте (только checkpoint, без пропуска по БД).",
    )
    p.add_argument(
        "--projects-file",
        type=Path,
        default=None,
        help=(
            "CSV/XLSX каталог всех проектов Harvest (Client, Project, Project Code, Currency). "
            "Нужен для проектов без записей времени (time report содержит только проекты с часами)."
        ),
    )
    g = p.add_mutually_exclusive_group(required=False)
    g.add_argument("--dry-run", action="store_true", help="Только статистика, без записи.")
    g.add_argument("--execute", action="store_true", help="Записать в БД.")
    args = p.parse_args()

    if args.progress:
        pass
    elif args.dry_run == args.execute:
        p.error("укажите ровно одно: --dry-run или --execute (или --progress)")

    if args.batch_size < 0:
        p.error("--batch-size must be >= 0")

    harvest_file = _resolve_harvest_file(args.file)
    if harvest_file is None:
        print(f"Файл не найден: {args.file}")
        print(f"Ожидаемые имена: {HARVEST_CSV_NAME} или {HARVEST_XLSX_NAME}")
        print("Проверьте пути:")
        for c in _harvest_file_candidates(args.file):
            print(f"  - {c}")
        print(
            "\nКонтейнер time_tracking — скопируйте CSV с хоста:\n"
            f"  docker cp timetrackinck/{HARVEST_CSV_NAME} "
            "$(docker compose ps -q time_tracking):/tmp/" + HARVEST_CSV_NAME
        )
        return 1

    projects_catalog_file: Path | None = None
    if args.projects_file is not None:
        projects_catalog_file = _resolve_projects_catalog_file(
            args.projects_file,
            time_report=harvest_file,
        )
        if projects_catalog_file is None:
            print(f"Каталог проектов не найден: {args.projects_file}")
            print("Проверьте пути:")
            for c in _projects_catalog_candidates(args.projects_file, harvest_file):
                print(f"  - {c}")
            print(
                "\nФайл нужно экспортировать из Harvest (Reports → Projects) и скопировать в контейнер:\n"
                f"  docker cp harvest_projects_all.csv "
                "$(docker compose ps -q time_tracking):/app/harvest_projects_all.csv"
            )
            print(
                "\nБез каталога импортируются только проекты с часами в time report (~749). "
                "Запуск без --projects-file тоже допустим."
            )
            fallback_catalog = (harvest_file.parent / args.projects_file.name).resolve()
            try:
                rows_for_catalog = _load_rows(harvest_file)
                generated = _write_projects_catalog_from_time_rows(
                    rows_for_catalog, fallback_catalog
                )
                projects_catalog_file = fallback_catalog
                print(
                    f"\nСгенерирован временный каталог: {fallback_catalog} "
                    f"({generated} проект(ов) из time report)."
                )
                print(
                    "ВНИМАНИЕ: это только проекты с часами. "
                    "Для полного 1:1 (все 904) замените файл экспортом Harvest Projects."
                )
            except Exception as e:
                print(f"Не удалось сгенерировать временный каталог: {e}")
                return 1

    from sqlalchemy.exc import IntegrityError

    database_url = _resolve_database_url(args.database_url or None)
    auth_db_url = (args.auth_db_url or os.environ.get("AUTH_DATABASE_URL") or "").strip()
    _configure_database_url(database_url)
    try:
        return asyncio.run(
            _run(
                path=harvest_file,
                execute=args.execute,
                database_url=database_url,
                auth_db_url=auth_db_url,
                replace=args.replace,
                team_only=args.team_only,
                batch_size=args.batch_size,
                checkpoint_file=args.checkpoint,
                reset_checkpoint=args.reset_checkpoint,
                progress_only=args.progress,
                client_filter=(args.client or "").strip() or None,
                project_filter=(args.project or "").strip() or None,
                project_index=args.project_index if args.project_index > 0 else None,
                sync_from_db=not args.no_sync_from_db,
                projects_file=projects_catalog_file,
            )
        )
    except KeyboardInterrupt:
        ckpt = args.checkpoint or _default_checkpoint_path(harvest_file)
        print(
            "\n\nОстановлено пользователем (Ctrl+C). "
            "Это не ошибка импорта."
        )
        if args.batch_size > 0 and ckpt.is_file():
            print(f"Прогресс сохранён в: {ckpt}")
            print("Продолжить: та же команда с --execute --batch-size 1")
            print("Статус:   добавьте --progress")
        return 130
    except IntegrityError as exc:
        print("\nОШИБКА БД: импорт прерван, текущий пакет не сохранён в checkpoint.")
        print(f"  {getattr(exc.orig, 'detail', exc) or exc}")
        print(
            "\nЗапустите ту же команду снова — продолжит с того же проекта "
            "(уже загруженные пропускаются автоматически)."
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
